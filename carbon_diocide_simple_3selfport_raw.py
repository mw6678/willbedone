import csv
from datetime import datetime
import sys
import os
import serial
import time
import re
import traceback
import serial.tools.list_ports
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, 
    QHBoxLayout, QLabel, QFrame, QComboBox, 
    QPushButton, QDialog, QMessageBox, QDesktopWidget, QMenu, QAction)
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QPoint
from PyQt5.QtGui import QFont, QColor
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# ==========================================
# [사용자 설정 영역]
# ==========================================
BAUD_RATE = 9600

# 센서별 2점 보정 파라미터 {센서index: (기울기 Slope, 절편 Offset)}
SENSOR_CALIB_PARAMS = {
    0: (1.0000, +0.0),  
    1: (1.084081, +0.0),  
    2: (1.00849, +0.0)   
}

# 센서 허용 측정 범위
CO2_MIN = 0
CO2_MAX = 30000

# 화면 표시용 이동평균 개수
SMOOTHING_WINDOW = 1


# ==========================================
# [포트 선택 다이얼로그]
# ==========================================
class PortSelectionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("시리얼 포트 설정")
        self.setFixedSize(340, 260)
        self.combos = []
        self.initUI()

    def initUI(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        title = QLabel("센서 포트 지정")
        title.setFont(QFont("Malgun Gothic", 12, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        available_ports = [port.device for port in serial.tools.list_ports.comports()]
        if not available_ports:
            available_ports = ["포트 없음"]

        for i in range(3):
            row = QHBoxLayout()
            lbl = QLabel(f"센서 {i+1}:")
            lbl.setFont(QFont("Malgun Gothic", 10))
            combo = QComboBox()
            combo.addItem("선택 안 함")
            for p in available_ports:
                if p != "포트 없음":
                    combo.addItem(p)
            self.combos.append(combo)
            row.addWidget(lbl)
            row.addWidget(combo)
            layout.addLayout(row)

        btn_start = QPushButton("모니터링 시작")
        btn_start.setFont(QFont("Malgun Gothic", 10, QFont.Bold))
        btn_start.setStyleSheet("padding: 8px; background-color: #007bff; color: white; border-radius: 4px;")
        btn_start.clicked.connect(self.on_start)
        layout.addWidget(btn_start)

    def on_start(self):
        selected = self.get_selected_ports()
        if not selected:
            QMessageBox.warning(self, "경고", "최소 1개 이상의 포트를 선택해야 합니다.")
            return
        self.accept()

    def get_selected_ports(self):
        selected = []
        for combo in self.combos:
            text = combo.currentText()
            if text not in ["선택 안 함", "포트 없음"]:
                selected.append(text)
        return selected


# ==========================================
# [백그라운드 개별 통신 스레드]
# ==========================================
class SerialThread(QThread):
    data_signal = pyqtSignal(int, int) 
    error_signal = pyqtSignal(int, str) 
    base_co2_signal = pyqtSignal(float)

    def __init__(self, port_name, sensor_index): 
        super().__init__()
        self.current_port = port_name  
        self.sensor_index = sensor_index
        self.smoothing_window = SMOOTHING_WINDOW
        self.data_buffer = []
        self.minute_data_buffer = [] 
        self.last_recorded_status = "정상"

    def parse_data(self, raw_data):
        try:
            raw_data = raw_data.strip()
            if not raw_data:
                return None

            raw_co2 = None
            
            # 특정 패턴 (A/B/C 등 구분 데이터 처리)
            if "A:" in raw_data or "B:" in raw_data or "C:" in raw_data:
                prefix = f"{chr(65 + self.sensor_index)}:"
                if prefix in raw_data:
                    match = re.search(f"{prefix}\\s*([-+]?\\d*\\.?\\d+)", raw_data)
                    if match:
                        raw_co2 = float(match.group(1))
            elif "CO2" in raw_data.upper():
                match = re.search(r'CO2\s*[:=]?\s*([-+]?\d*\.?\d+)', raw_data, re.IGNORECASE)
                if match:
                    raw_co2 = float(match.group(1))
            else:
                numbers = re.findall(r"[-+]?\d*\.?\d+", raw_data)
                if len(numbers) == 1:
                    raw_co2 = float(numbers[0])

            if raw_co2 is None:
                return None

            if raw_co2 < CO2_MIN or raw_co2 > CO2_MAX:
                return "OUT_OF_RANGE"

            return raw_co2
        except Exception:
            return None

    def safe_sleep(self, ms): 
        steps = max(1, ms // 50)
        for _ in range(steps):
            if self.isInterruptionRequested():
                return  
            QThread.msleep(50)

    def write_log(self, message):
        log_dir = os.path.join(BASE_DIR, "Logs")
        os.makedirs(log_dir, exist_ok=True)
        log_file = os.path.join(log_dir, f"error_log_Sensor{self.sensor_index + 1}.txt")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        try:
            with open(log_file, "a", encoding="utf-8") as f:
                f.write(f"{timestamp} | 포트: {self.current_port} | 사유: {message}\n")
        except Exception:
            pass

    def cleanup_old_logs(self):
        retention_days = 30
        threshold_time = time.time() - (retention_days * 24 * 60 * 60)
        directories = ["CSV_Logs", "Excel_Logs", "Logs"]

        for dir_name in directories:
            target_dir = os.path.join(BASE_DIR, dir_name)
            if not os.path.exists(target_dir):
                continue
            for filename in os.listdir(target_dir):
                file_path = os.path.join(target_dir, filename)
                if os.path.isfile(file_path):
                    try:
                        if os.path.getmtime(file_path) < threshold_time:
                            os.remove(file_path)
                    except Exception:
                        pass

    def save_excel_log(self, today_str, time_str, record_value):
        try:
            save_dir = os.path.join(BASE_DIR, "Excel_Logs")
            os.makedirs(save_dir, exist_ok=True)
            file_path = os.path.join(save_dir, f"CO2_log_{today_str}_Sensor{self.sensor_index + 1}.xlsx")

            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )

            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.title = "측정 데이터"

                headers = ["측정일자", "측정시간", "센서번호", "포트", "CO2 1분 평균(ppm)"]
                ws.append(headers)

                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                ws.row_dimensions[1].height = 25
            else:
                wb = load_workbook(file_path)
                ws = wb.active

            row_data = [today_str, time_str, self.sensor_index + 1, self.current_port, record_value]
            ws.append(row_data)

            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20

            data_font = Font(name="맑은 고딕", size=10)
            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")

            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.font = data_font
                cell.border = thin_border

                if col_num in [1, 2, 3, 4]:
                    cell.alignment = center_align
                else:
                    if isinstance(val, (int, float)):
                        cell.alignment = right_align
                        cell.number_format = '#,##0'
                    else:
                        cell.alignment = center_align

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row == 1:
                        val_str = str(cell.value or '')
                        max_len = max(max_len, len(val_str.encode('utf-8')))
                    else:
                        val_str = str(cell.value or '')
                        max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

            wb.save(file_path)

        except Exception as e:
            self.write_log(f"Excel 저장 오류: {str(e)}")

    def run(self):
        self.cleanup_old_logs()

        is_connected = False 
        is_no_data_error_sent = False 

        last_saved_minute = time.localtime().tm_min
        last_ui_update_time = 0.0

        while not self.isInterruptionRequested():
            ser = None
            try:
                ser = serial.Serial(self.current_port, BAUD_RATE, timeout=0.1)
                
                if not is_connected: 
                    self.write_log("통신 연결 성공")
                    is_connected = True
                    self.data_buffer.clear()
                    self.minute_data_buffer.clear()
                    self.last_recorded_status = "정상"
                    ser.reset_input_buffer()

                last_data_time = time.time()

                while ser.is_open and not self.isInterruptionRequested():
                    calibrated_co2_value = None
                    current_time = time.time()

                    # 1. 시리얼 수신 데이터 처리
                    while ser.in_waiting > 0:
                        raw_bytes = ser.readline()
                        if not raw_bytes: 
                            continue

                        raw_data = raw_bytes.decode("utf-8", errors="ignore").strip()
                        parsed = self.parse_data(raw_data)
                        
                        if parsed == "OUT_OF_RANGE":
                            err_msg = "범위 초과 (위험)"
                            self.error_signal.emit(self.sensor_index, err_msg)
                            self.write_log(err_msg)
                            self.last_recorded_status = err_msg
                            last_data_time = current_time
                            continue

                        if parsed is not None:
                            slope, offset = SENSOR_CALIB_PARAMS.get(self.sensor_index, (1.0, 0.0))
                            calibrated_co2_value = (parsed * slope) + offset
                            last_data_time = current_time
                            self.last_recorded_status = "정상"

                    # 2. 통신 끊김 체크 (30초 타임아웃)
                    if current_time - last_data_time >= 30.0:
                        if not is_no_data_error_sent:
                            err_msg = "데이터 없음 (30초 초과)"
                            self.error_signal.emit(self.sensor_index, "데이터 없음")
                            self.write_log(err_msg)
                            self.last_recorded_status = "데이터 없음"
                            is_no_data_error_sent = True
                            self.data_buffer.clear()

                    # 3. 데이터 버퍼 및 UI 갱신 (1초 주기로 제한)
                    elif calibrated_co2_value is not None:
                        is_no_data_error_sent = False
                        self.data_buffer.append(calibrated_co2_value)
                        self.minute_data_buffer.append(calibrated_co2_value)
                        
                        if len(self.data_buffer) > self.smoothing_window:
                            self.data_buffer.pop(0)

                        if current_time - last_ui_update_time >= 1.0:
                            last_ui_update_time = current_time
                            smoothed_co2 = int(sum(self.data_buffer) / len(self.data_buffer))
                            self.data_signal.emit(self.sensor_index, smoothed_co2)

                    # 4. 1분 단위 CSV 및 Excel 저장 로직
                    now_struct = time.localtime(current_time)
                    if now_struct.tm_min != last_saved_minute:
                        last_saved_minute = now_struct.tm_min
                        
                        now_dt = datetime.now()
                        today_str = now_dt.strftime("%Y-%m-%d")
                        time_str = now_dt.strftime("%H:%M:00")

                        if self.minute_data_buffer:
                            record_value = int(sum(self.minute_data_buffer) / len(self.minute_data_buffer))
                            self.minute_data_buffer.clear()
                        else:
                            record_value = f"Error: {self.last_recorded_status}"

                        # CSV 저장
                        save_dir = os.path.join(BASE_DIR, "CSV_Logs")
                        os.makedirs(save_dir, exist_ok=True)
                        file_path = os.path.join(save_dir, f"CO2_log_{today_str}_Sensor{self.sensor_index + 1}.csv")
                        file_exists = os.path.exists(file_path)

                        try:
                            with open(file_path, mode='a', newline='', encoding='utf-8-sig') as f:
                                writer = csv.writer(f)
                                if not file_exists:
                                    writer.writerow(["측정일자", "측정시간", "센서번호", "포트", "Co2 1분 평균(ppm)"])
                                writer.writerow([today_str, time_str, self.sensor_index + 1, self.current_port, record_value])
                        except Exception as e:
                            self.write_log(f"CSV 저장 오류: {str(e)}")

                        # Excel 저장
                        self.save_excel_log(today_str, time_str, record_value)

                    self.safe_sleep(50)

            except serial.SerialException as se:
                if is_connected: 
                    self.write_log(f"시리얼 연결 실패/끊김: {se}")
                    is_connected = False
                self.data_buffer.clear()
                self.error_signal.emit(self.sensor_index, "연결 실패/끊김")
                self.last_recorded_status = "연결 끊김"
            except Exception as e:
                if is_connected:
                    self.write_log(f"시스템 예외 오류: {traceback.format_exc()}")
                    is_connected = False
                self.data_buffer.clear()
                self.error_signal.emit(self.sensor_index, "시스템 오류")
                self.last_recorded_status = "시스템 오류"
            finally:
                if ser is not None:
                    try:
                        if ser.is_open:
                            ser.close()
                    except Exception:
                        pass
                self.safe_sleep(2000)


# ==========================================
# [CO2 레벨 표시 위젯]
# ==========================================
class CO2LevelWidget(QWidget):
    LEVELS = [
        {"name": "좋음", "min": 1, "max": 500, "color": "#28A745"},        
        {"name": "보통", "min": 501, "max": 1000, "color": "#FFD700"},        
        {"name": "민감군", "min": 1001, "max": 3000, "color": "#FD7E14"},     
        {"name": "나쁨", "min": 3001, "max": 5000, "color": "#DC3545"},      
        {"name": "매우 나쁨", "min": 5001, "max": 10000, "color": "#800080"}, 
        {"name": "위험", "min": 10001, "max": 30000, "color": "#795548"},      
    ]

    def __init__(self, sensor_index, title="센서", parent=None):
        super().__init__(parent)
        self.sensor_index = sensor_index
        self.title = title
        self.current_value = 400 
        self.initUI()

    def initUI(self):
        self.setStyleSheet("background-color: transparent; border: none;")
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(5)

        self.title_label = QLabel(self.title)
        self.title_label.setFont(QFont("Malgun Gothic", 16, QFont.Bold))
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("color: black; border: none;")
        main_layout.addWidget(self.title_label)
        
        main_layout.addSpacing(5)

        value_layout = QHBoxLayout()
        value_layout.setAlignment(Qt.AlignCenter)
        value_layout.setSpacing(2)

        self.co2_value_label = QLabel("----")
        self.co2_value_label.setFont(QFont("Arial", 42, QFont.Bold))
        self.co2_value_label.setStyleSheet("color: black; border: none;")
        value_layout.addWidget(self.co2_value_label)

        ppm_label = QLabel("ppm")
        ppm_label.setFont(QFont("Arial", 16, QFont.Bold))
        ppm_label.setStyleSheet("color: black; margin-bottom: 8px; border: none;") 
        ppm_label.setAlignment(Qt.AlignBottom)
        value_layout.addWidget(ppm_label)

        main_layout.addLayout(value_layout)
        main_layout.addSpacing(10)

        self.arrow_container = QWidget()
        self.arrow_container.setStyleSheet("border: none;")
        self.arrow_container.setFixedHeight(20)
        
        self.arrow_label = QLabel("▼")
        self.arrow_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.arrow_label.setStyleSheet("color: black; border: none;")
        self.arrow_label.setAlignment(Qt.AlignCenter)
        self.arrow_label.setFixedWidth(20)
        self.arrow_label.setParent(self.arrow_container)
        self.arrow_label.hide()
        main_layout.addWidget(self.arrow_container)

        level_bar_frame = QFrame()
        level_bar_frame.setStyleSheet("border: none;")
        self.level_bar_layout = QHBoxLayout(level_bar_frame)
        self.level_bar_layout.setContentsMargins(0, 0, 0, 0)
        self.level_bar_layout.setSpacing(2) 

        self.level_bars = []
        for i, level in enumerate(self.LEVELS):
            bar = QFrame()
            bar.setFixedHeight(14)
            
            border_radius = ""
            if i == 0:
                border_radius = "border-top-left-radius: 7px; border-bottom-left-radius: 7px;"
            elif i == len(self.LEVELS) - 1:
                border_radius = "border-top-right-radius: 7px; border-bottom-right-radius: 7px;"
                
            bar.setStyleSheet(f"background-color: {level['color']}; {border_radius}")
            self.level_bars.append(bar)
            self.level_bar_layout.addWidget(bar)

        main_layout.addWidget(level_bar_frame)
        main_layout.addSpacing(15)

        self.status_text_label = QLabel("대기 중...")
        self.status_text_label.setFont(QFont("Malgun Gothic", 16, QFont.Bold))
        self.status_text_label.setAlignment(Qt.AlignCenter)
        self.status_text_label.setFixedHeight(45)
        self.status_text_label.setStyleSheet("background-color: #E0E0E0; border-radius: 8px; color: gray;")
        
        main_layout.addWidget(self.status_text_label)
        main_layout.addStretch(1)

    def update_co2(self, value):
        self.arrow_label.show()         
        self.current_value = value
        self.co2_value_label.setText(str(value))

        current_level = None
        for level in self.LEVELS:
            if value <= level['max']:
                current_level = level
                break
        
        if value < self.LEVELS[0]['min']: 
            current_level = self.LEVELS[0]
        elif value > self.LEVELS[-1]['max']: 
            current_level = self.LEVELS[-1]

        if current_level:
            self.status_text_label.setText(current_level['name'])
            color = current_level['color']
            text_color = "black" if current_level['name'] == "보통" else "white"
            self.status_text_label.setStyleSheet(f"background-color: {color}; border-radius: 8px; color: {text_color};")

        self.update_arrow_position()

    def set_error_state(self, msg):
        self.co2_value_label.setText("----")
        self.status_text_label.setText(msg)
        self.status_text_label.setStyleSheet("background-color: #FFCDD2; border-radius: 8px; color: #B71C1C;")
        
        self.current_value = self.LEVELS[0]['min']
        self.update_arrow_position()
        self.arrow_label.hide()

    def update_arrow_position(self):
        if not self.level_bars or self.arrow_container.width() <= 0:
            return 

        target_bar_width = self.level_bars[0].geometry().width()
        if target_bar_width <= 0:
            return

        value = self.current_value
        
        level_index = 0
        for i, level in enumerate(self.LEVELS):
            if value <= level['max']:
                level_index = i
                break
        else:
            level_index = len(self.LEVELS) - 1

        current_level = self.LEVELS[level_index]
        level_min = current_level['min']
        level_max = current_level['max']

        if value <= level_min: ratio = 0.0
        elif value >= level_max: ratio = 1.0
        else: ratio = (value - level_min) / (level_max - level_min)

        target_bar = self.level_bars[level_index]
        bar_x = target_bar.geometry().x()
        bar_width = target_bar.geometry().width()

        target_x = bar_x + (bar_width * ratio)
        arrow_x = int(target_x - (self.arrow_label.width() / 2))
        
        max_x = self.arrow_container.width() - self.arrow_label.width()
        arrow_x = max(0, min(arrow_x, max_x))
        
        arrow_y = self.arrow_container.height() - self.arrow_label.height()

        self.arrow_label.move(QPoint(arrow_x, int(arrow_y)))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_arrow_position()


# ==========================================
# [메인 GUI 클래스]
# ==========================================
class CO2MonitorApp(QMainWindow):
    def __init__(self, target_ports):
        super().__init__()
        self.target_ports = target_ports
        self.threads = []
        self.dragPos = QPoint()
        self.is_always_on_top = False
        self.initUI()
        self.start_threads()

    def initUI(self):
        self.setWindowTitle("이산화탄소 다중 모니터링")
        
        num_ports = len(self.target_ports)
        window_width = max(280, num_ports * 260)
        self.resize(window_width, 360) 
        
        self.setStyleSheet("QMainWindow { background-color: #E9ECEF; border: none; }") 
        self.setWindowFlags(Qt.FramelessWindowHint)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15) 

        self.widgets = []
        for i, port in enumerate(self.target_ports):
            widget = CO2LevelWidget(sensor_index=i, title=f"센서 {i+1} ({port})") 
            self.widgets.append(widget)
            main_layout.addWidget(widget)

    def start_threads(self):
        for i, port_name in enumerate(self.target_ports): 
            thread = SerialThread(port_name, i) 
            thread.data_signal.connect(self.update_data)
            thread.error_signal.connect(self.handle_error) 
            self.threads.append(thread)
            
        for thread in self.threads:
            thread.start()

    def update_data(self, sensor_index, co2_value):
        if 0 <= sensor_index < len(self.widgets):
            self.widgets[sensor_index].update_co2(co2_value)

    def handle_error(self, sensor_index, error_msg):
        if 0 <= sensor_index < len(self.widgets):
            self.widgets[sensor_index].set_error_state(error_msg)

    def contextMenuEvent(self, event):
        menu = QMenu(self)
        
        toggle_top_action = QAction("최상단 고정 켜기/끄기", self)
        toggle_top_action.triggered.connect(self.toggle_always_on_top)
        menu.addAction(toggle_top_action)

        capture_action = QAction("화면 캡처", self)
        capture_action.triggered.connect(self.capture_screen)
        menu.addAction(capture_action)

        exit_action = QAction("종료", self)
        exit_action.triggered.connect(self.close)
        menu.addAction(exit_action)

        menu.exec_(event.globalPos())

    def toggle_always_on_top(self):
        self.is_always_on_top = not self.is_always_on_top
        if self.is_always_on_top:
            self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        else:
            self.setWindowFlags(self.windowFlags() & ~Qt.WindowStaysOnTopHint)
        self.show()

    def capture_screen(self):
        screen = QApplication.primaryScreen()
        screenshot = screen.grabWindow(self.winId())
        save_dir = os.path.join(BASE_DIR, "Captures")
        os.makedirs(save_dir, exist_ok=True)
        filename = datetime.now().strftime("capture_%Y%m%d_%H%M%S.png")
        file_path = os.path.join(save_dir, filename)
        screenshot.save(file_path, "PNG")
        QMessageBox.information(self, "캡처 완료", f"화면이 저장되었습니다:\n{file_path}")

    def closeEvent(self, event):
        for thread in self.threads:
            thread.requestInterruption()
            thread.wait(1000)
        event.accept()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.dragPos = event.globalPos() - self.frameGeometry().topLeft()

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton:
            self.move(event.globalPos() - self.dragPos)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    setup_dialog = PortSelectionDialog()
    
    if setup_dialog.exec_() == QDialog.Accepted:
        selected_ports = setup_dialog.get_selected_ports()
        window = CO2MonitorApp(selected_ports)
        window.show()
        sys.exit(app.exec_())