import sys
import json
import os
from pathlib import Path
import urllib.request
import urllib.parse
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit, 
    QPushButton, QComboBox, QRadioButton, QButtonGroup, 
    QTableWidget, QTableWidgetItem, QHeaderView, 
    QVBoxLayout, QHBoxLayout, QGridLayout, QGroupBox, 
    QMessageBox, QDialog, QFrame, QScrollArea, QFileDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QFont as QGuiFont, QColor, QPalette

APP_NAME = "주유비 계산기 v2.0"
APP_DIR = Path(os.getenv("APPDATA", Path.home())) / "FuelCalculator"
CONFIG_FILE = APP_DIR / "config.json"
OPINET_API_URL = "https://www.opinet.co.kr/api/avgAllPrice.do"


class ConfigManager:
    """설정 파일(API 키, 차량 목록) 관리 클래스"""
    def __init__(self):
        self.config = {"api_key": "", "vehicles": []}
        self.load()

    def load(self):
        try:
            if CONFIG_FILE.exists():
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, dict):
                        self.config.update(data)
        except Exception:
            self.config = {"api_key": "", "vehicles": []}

    def save(self):
        try:
            APP_DIR.mkdir(parents=True, exist_ok=True)
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(self.config, f, ensure_ascii=False, indent=4)
            return True
        except Exception as e:
            print(f"설정 저장 오류: {e}")
            return False

    def get_api_key(self):
        # 여기에 본인의 오피넷 API 키를 직접 넣어두면 앱 실행 시 기본으로 사용됩니다.
        # 예: return self.config.get("api_key") or "여기에_본인_API키_입력"
        return self.config.get("api_key") or "여기에_본인_오피넷_API_키_입력"

    def set_api_key(self, api_key):
        self.config["api_key"] = api_key
        self.save()

    def get_vehicles(self):
        vehicles = self.config.get("vehicles", [])
        return vehicles if isinstance(vehicles, list) else []

    def add_vehicle(self, name, fuel_type, efficiency):
        vehicle = {"name": name, "fuel_type": fuel_type, "efficiency": efficiency}
        vehicles = self.get_vehicles()
        vehicles.append(vehicle)
        self.config["vehicles"] = vehicles
        self.save()

    def delete_vehicle(self, index):
        vehicles = self.get_vehicles()
        if 0 <= index < len(vehicles):
            del vehicles[index]
            self.config["vehicles"] = vehicles
            self.save()


class VehicleManagerDialog(QDialog):
    """차량 관리 대화상자"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.setWindowTitle("차량 관리")
        self.resize(500, 450)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)

        list_group = QGroupBox("저장된 차량 목록")
        list_layout = QVBoxLayout(list_group)
        
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["차량명", "유종", "연비 (km/L)"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        list_layout.addWidget(self.table)
        layout.addWidget(list_group)

        add_group = QGroupBox("새 차량 추가")
        add_layout = QGridLayout(add_group)

        add_layout.addWidget(QLabel("차량명"), 0, 0)
        self.name_input = QLineEdit()
        add_layout.addWidget(self.name_input, 0, 1)

        add_layout.addWidget(QLabel("유종"), 0, 2)
        self.fuel_combo = QComboBox()
        self.fuel_combo.addItems(["휘발유", "경유", "LPG"])
        add_layout.addWidget(self.fuel_combo, 0, 3)

        add_layout.addWidget(QLabel("연비"), 1, 0)
        self.efficiency_input = QLineEdit()
        add_layout.addWidget(self.efficiency_input, 1, 1)
        add_layout.addWidget(QLabel("km/L"), 1, 2)

        add_btn = QPushButton("차량 추가")
        add_btn.clicked.connect(self.add_vehicle)
        add_layout.addWidget(add_btn, 1, 3)

        layout.addWidget(add_group)

        del_btn = QPushButton("선택 차량 삭제")
        del_btn.setStyleSheet("background-color: #ffebee; color: #c62828; font-weight: bold;")
        del_btn.clicked.connect(self.delete_vehicle)
        layout.addWidget(del_btn)

        self.refresh_table()

    def refresh_table(self):
        vehicles = self.parent_app.config.get_vehicles()
        self.table.setRowCount(len(vehicles))
        for i, v in enumerate(vehicles):
            self.table.setItem(i, 0, QTableWidgetItem(v["name"]))
            self.table.setItem(i, 1, QTableWidgetItem(v["fuel_type"]))
            self.table.setItem(i, 2, QTableWidgetItem(str(v["efficiency"])))

    def add_vehicle(self):
        name = self.name_input.text().strip()
        fuel = self.fuel_combo.currentText()
        eff_text = self.efficiency_input.text().strip()

        if not name:
            QMessageBox.warning(self, "입력 오류", "차량명을 입력해주세요.")
            return

        try:
            efficiency = float(eff_text)
            if efficiency <= 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "입력 오류", "유효한 연비 숫자를 입력해주세요.")
            return

        self.parent_app.config.add_vehicle(name, fuel, efficiency)
        self.refresh_table()
        self.parent_app.load_vehicles()
        self.name_input.clear()
        self.efficiency_input.clear()

    def delete_vehicle(self):
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, "선택 오류", "삭제할 차량을 선택해주세요.")
            return

        vehicles = self.parent_app.config.get_vehicles()
        target = vehicles[selected]

        confirm = QMessageBox.question(
            self, "차량 삭제", f"'{target['name']}' 차량을 삭제하시겠습니까?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if confirm == QMessageBox.StandardButton.Yes:
            self.parent_app.config.delete_vehicle(selected)
            self.refresh_table()
            self.parent_app.load_vehicles()


class ApiKeyDialog(QDialog):
    """API 키 설정 대화상자"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.setWindowTitle("오피넷 API 키 설정")
        self.resize(400, 180)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)

        layout.addWidget(QLabel("<b>오피넷 API 인증키 입력</b>"))
        
        self.key_input = QLineEdit()
        self.key_input.setText(self.parent_app.config.get_api_key())
        self.key_input.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(self.key_input)

        save_btn = QPushButton("저장")
        save_btn.clicked.connect(self.save_key)
        layout.addWidget(save_btn)

    def save_key(self):
        key = self.key_input.text().strip()
        if not key:
            QMessageBox.warning(self, "입력 오류", "API 키를 입력해주세요.")
            return
        self.parent_app.config.set_api_key(key)
        self.parent_app.api_status_label.setText("● API 키 저장됨")
        self.parent_app.api_status_label.setStyleSheet("color: #2E7D32; font-size: 11px;")
        self.accept()


class FuelCalculatorWindow(QMainWindow):
    """메인 창"""
    def __init__(self):
        super().__init__()
        self.config = ConfigManager()
        self.setWindowTitle(APP_NAME)
        self.resize(650, 920)
        self.setMinimumSize(600, 800)
        
        self.init_ui()
        self.load_vehicles()

        if self.config.get_api_key():
            self.api_status_label.setText("● API 키 저장됨")
            self.api_status_label.setStyleSheet("color: #2E7D32; font-size: 11px;")

    def init_ui(self):
        self.setStyleSheet("""
            QMainWindow { background-color: #F4F6F8; }
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 1px solid #D0D7DE;
                border-radius: 6px;
                margin-top: 10px;
                background-color: #FFFFFF;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
                color: #333333;
            }
            QLabel { font-size: 12px; color: #333333; }
            QLineEdit, QComboBox {
                font-size: 12px;
                padding: 6px;
                border: 1px solid #CCC;
                border-radius: 4px;
                background-color: #FAFAFA;
            }
            QPushButton {
                font-size: 12px;
                padding: 6px 12px;
                border: 1px solid #CCC;
                border-radius: 4px;
                background-color: #EFEFEF;
            }
            QPushButton:hover { background-color: #E0E0E0; }
        """)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # 1. 헤더 영역
        header_frame = QFrame()
        header_frame.setStyleSheet("background-color: #163A5F;")
        header_layout = QVBoxLayout(header_frame)
        header_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_layout.setContentsMargins(10, 20, 10, 20)

        title_label = QLabel("🚗 주유비 계산기")
        title_label.setStyleSheet("font-size: 22px; font-weight: bold; color: white;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(title_label)

        sub_label = QLabel("주행거리 · 연비 · 전국 평균 유가를 이용한 스마트 주유비 계산")
        sub_label.setStyleSheet("font-size: 11px; color: #DCE7F2;")
        sub_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        header_layout.addWidget(sub_label)

        main_layout.addWidget(header_frame)

        # 2. 스크롤 영역
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setContentsMargins(20, 15, 20, 15)
        content_layout.setSpacing(12)

        # --- 차량 정보 섹션 ---
        vehicle_group = QGroupBox(" 차량 정보 ")
        v_layout = QGridLayout(vehicle_group)
        v_layout.addWidget(QLabel("차량 선택"), 0, 0)
        
        self.vehicle_combo = QComboBox()
        self.vehicle_combo.currentIndexChanged.connect(self.vehicle_selected)
        v_layout.addWidget(self.vehicle_combo, 0, 1)

        vehicle_mgr_btn = QPushButton("차량 관리")
        vehicle_mgr_btn.clicked.connect(self.open_vehicle_manager)
        v_layout.addWidget(vehicle_mgr_btn, 0, 2)
        content_layout.addWidget(vehicle_group)

        # --- 주행 정보 섹션 ---
        driving_group = QGroupBox(" 주행 정보 ")
        d_layout = QGridLayout(driving_group)
        
        d_layout.addWidget(QLabel("주행거리"), 0, 0)
        self.distance_input = QLineEdit()
        d_layout.addWidget(self.distance_input, 0, 1)
        d_layout.addWidget(QLabel("km"), 0, 2)

        d_layout.addWidget(QLabel("연비"), 1, 0)
        self.efficiency_input = QLineEdit()
        d_layout.addWidget(self.efficiency_input, 1, 1)
        d_layout.addWidget(QLabel("km/L"), 1, 2)

        d_layout.addWidget(QLabel("주행방식"), 2, 0)
        trip_layout = QHBoxLayout()
        self.trip_one_way = QRadioButton("편도")
        self.trip_one_way.setChecked(True)
        self.trip_round_trip = QRadioButton("왕복")
        self.trip_group = QButtonGroup(self)
        self.trip_group.addButton(self.trip_one_way)
        self.trip_group.addButton(self.trip_round_trip)
        trip_layout.addWidget(self.trip_one_way)
        trip_layout.addWidget(self.trip_round_trip)
        trip_layout.addStretch()
        d_layout.addLayout(trip_layout, 2, 1, 1, 2)
        
        content_layout.addWidget(driving_group)

        # --- 유가 정보 섹션 ---
        price_group = QGroupBox(" 유가 정보 ")
        p_layout = QGridLayout(price_group)

        p_layout.addWidget(QLabel("유종"), 0, 0)
        self.fuel_combo = QComboBox()
        self.fuel_combo.addItems(["휘발유", "경유", "LPG"])
        self.fuel_combo.currentIndexChanged.connect(self.update_price)
        p_layout.addWidget(self.fuel_combo, 0, 1)

        p_layout.addWidget(QLabel("적용 유가"), 1, 0)
        self.price_input = QLineEdit()
        p_layout.addWidget(self.price_input, 1, 1)
        p_layout.addWidget(QLabel("원/L"), 1, 2)

        fetch_price_btn = QPushButton("전국 평균 유가 조회")
        fetch_price_btn.clicked.connect(self.update_price)
        p_layout.addWidget(fetch_price_btn, 2, 1)

        self.api_status_label = QLabel("● API 키 없음")
        self.api_status_label.setStyleSheet("color: #777777; font-size: 11px;")
        p_layout.addWidget(self.api_status_label, 2, 2)

        self.price_source_label = QLabel("유가 정보를 조회해주세요.")
        self.price_source_label.setStyleSheet("color: #555555; font-size: 11px;")
        p_layout.addWidget(self.price_source_label, 3, 0, 1, 3)

        self.price_date_label = QLabel("")
        self.price_date_label.setStyleSheet("color: #777777; font-size: 11px;")
        p_layout.addWidget(self.price_date_label, 4, 0, 1, 3)

        content_layout.addWidget(price_group)

        # --- 추가 비용 섹션 ---
        extra_group = QGroupBox(" 추가 비용 ")
        e_layout = QGridLayout(extra_group)
        
        e_layout.addWidget(QLabel("통행료"), 0, 0)
        self.toll_input = QLineEdit("0")
        e_layout.addWidget(self.toll_input, 0, 1)
        e_layout.addWidget(QLabel("원"), 0, 2)

        e_layout.addWidget(QLabel("주차비"), 1, 0)
        self.parking_input = QLineEdit("0")
        e_layout.addWidget(self.parking_input, 1, 1)
        e_layout.addWidget(QLabel("원"), 1, 2)

        e_layout.addWidget(QLabel("기타 비용"), 2, 0)
        self.other_input = QLineEdit("0")
        e_layout.addWidget(self.other_input, 2, 1)
        e_layout.addWidget(QLabel("원"), 2, 2)

        content_layout.addWidget(extra_group)

        # --- 계산/초기화 버튼 ---
        btn_layout = QHBoxLayout()
        calc_btn = QPushButton("계산하기")
        calc_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976D2; color: white; font-weight: bold; font-size: 14px; padding: 10px;
                border-radius: 6px;
            }
            QPushButton:hover { background-color: #115293; }
        """)
        calc_btn.clicked.connect(self.calculate)
        btn_layout.addWidget(calc_btn)

        reset_btn = QPushButton("초기화")
        reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #E0E0E0; color: #333333; font-size: 14px; padding: 10px;
                border-radius: 6px;
            }
            QPushButton:hover { background-color: #CCCCCC; }
        """)
        reset_btn.clicked.connect(self.reset_fields)
        btn_layout.addWidget(reset_btn)

        content_layout.addLayout(btn_layout)

        # --- 계산 결과 섹션 ---
        result_group = QGroupBox(" 계산 결과 ")
        r_layout = QGridLayout(result_group)

        self.res_distance = QLabel("-")
        self.res_fuel = QLabel("-")
        self.res_fuel_cost = QLabel("-")
        self.res_per_km = QLabel("-")
        self.res_per_100km = QLabel("-")
        self.res_extra = QLabel("-")
        self.res_total = QLabel("-")

        results_info = [
            ("총 주행거리", self.res_distance, "km"),
            ("예상 주유량", self.res_fuel, "L"),
            ("예상 주유비", self.res_fuel_cost, "원"),
            ("1 km당 연료비", self.res_per_km, "원"),
            ("100 km당 연료비", self.res_per_100km, "원"),
            ("추가 비용", self.res_extra, "원"),
        ]

        for idx, (label_text, lbl_val, unit) in enumerate(results_info):
            r_layout.addWidget(QLabel(label_text), idx, 0)
            lbl_val.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            lbl_val.setStyleSheet("font-weight: bold;")
            r_layout.addWidget(lbl_val, idx, 1)
            r_layout.addWidget(QLabel(unit), idx, 2)

        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        r_layout.addWidget(line, len(results_info), 0, 1, 3)

        total_title = QLabel("총 여행비")
        total_title.setStyleSheet("font-size: 14px; font-weight: bold;")
        r_layout.addWidget(total_title, len(results_info) + 1, 0)

        self.res_total.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.res_total.setStyleSheet("font-size: 18px; font-weight: bold; color: #D32F2F;")
        r_layout.addWidget(self.res_total, len(results_info) + 1, 1)

        total_unit = QLabel("원")
        total_unit.setStyleSheet("font-size: 13px; font-weight: bold;")
        r_layout.addWidget(total_unit, len(results_info) + 1, 2)

        content_layout.addWidget(result_group)

        # --- 엑셀 저장 버튼 추가 ---
        self.excel_save_btn = QPushButton("📊 엑셀로 저장하기")
        self.excel_save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2E7D32; color: white; font-weight: bold; font-size: 14px; padding: 12px;
                border-radius: 6px;
            }
            QPushButton:hover { background-color: #1B5E20; }
        """)
        self.excel_save_btn.clicked.connect(self.save_to_excel)
        content_layout.addWidget(self.excel_save_btn)

        notice_label = QLabel("※ 유가는 오피넷 전국 평균가격 기준이며 실제 주유가격과 차이가 있을 수 있습니다.")
        notice_label.setStyleSheet("color: #777777; font-size: 10px;")
        notice_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        content_layout.addWidget(notice_label)

        scroll_area.setWidget(content_widget)
        main_layout.addWidget(scroll_area)

    def load_vehicles(self):
        self.vehicle_combo.blockSignals(True)
        self.vehicle_combo.clear()
        vehicles = self.config.get_vehicles()
        for v in vehicles:
            self.vehicle_combo.addItem(v["name"], v)
        self.vehicle_combo.blockSignals(False)

        if vehicles:
            self.vehicle_combo.setCurrentIndex(0)
            self.vehicle_selected()

    def vehicle_selected(self):
        v_data = self.vehicle_combo.currentData()
        if v_data:
            index = self.fuel_combo.findText(v_data["fuel_type"])
            if index >= 0:
                self.fuel_combo.setCurrentIndex(index)
            self.efficiency_input.setText(str(v_data["efficiency"]))
            self.update_price()

    def open_vehicle_manager(self):
        dialog = VehicleManagerDialog(self)
        dialog.exec()

    def update_price(self):
        api_key = self.config.get_api_key().strip()
        if not api_key:
            if QMessageBox.question(self, "API 키 필요", "오피넷 API 키가 없습니다. 설정하시겠습니까?") == QMessageBox.StandardButton.Yes:
                apiKeyDialog = ApiKeyDialog(self)
                apiKeyDialog.exec()
            return

        fuel_type = self.fuel_combo.currentText()
        
        # 오피넷 API가 반환하는 실제 PRODNM 값과 매칭 (여기가 원인이었습니다)
        target_prodnm_list = {
            "휘발유": ["휘발유"],
            "경유": ["자동차용경유", "경유"],
            "LPG": ["자동차용부탄", "LPG", "부탄"]
        }.get(fuel_type, [fuel_type])

        try:
            params = {"out": "xml", "code": api_key}
            url = f"{OPINET_API_URL}?{urllib.parse.urlencode(params)}"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            
            with urllib.request.urlopen(req, timeout=10) as response:
                data = response.read()

            root = ET.fromstring(data)

            price = None
            date = None

            # 오피넷 XML에서 유종 매칭
            for oil_item in root.findall(".//OIL"):
                prodnm_el = oil_item.find("PRODNM")
                if prodnm_el is not None and prodnm_el.text in target_prodnm_list:
                    price_el = oil_item.find("PRICE")
                    date_el = oil_item.find("TRADE_DT")
                    
                    if price_el is not None and price_el.text:
                        price = float(price_el.text)
                    if date_el is not None and date_el.text:
                        date = date_el.text
                    break

            if price is None:
                raise ValueError(f"'{fuel_type}'에 해당하는 유가 정보를 API 응답에서 찾지 못했습니다.")

            self.price_input.setText(f"{price:.2f}")
            self.price_source_label.setText("출처: 오피넷 전국 평균가격")
            self.price_date_label.setText(f"기준일: {date}" if date else "기준일 정보 없음")

        except Exception as e:
            QMessageBox.critical(self, "유가 조회 실패", f"유가 정보를 가져오는 중 오류가 발생했습니다.\n{e}")
            self.price_input.setText(f"{price:.2f}")
            self.price_source_label.setText("출처: 오피넷 전국 평균가격")
            self.price_date_label.setText(f"기준일: {date}" if date else "기준일 정보 없음")

        except Exception as e:
            QMessageBox.critical(self, "유가 조회 실패", f"유가 정보를 가져오는 중 오류가 발생했습니다.\n{e}")

    def calculate(self):
        try:
            distance = float(self.distance_input.text().replace(",", "") or 0)
            efficiency = float(self.efficiency_input.text().replace(",", "") or 0)
            price = float(self.price_input.text().replace(",", "") or 0)
            toll = float(self.toll_input.text().replace(",", "") or 0)
            parking = float(self.parking_input.text().replace(",", "") or 0)
            other = float(self.other_input.text().replace(",", "") or 0)
        except ValueError:
            QMessageBox.warning(self, "입력 오류", "거리, 연비, 유가 및 비용에 올바른 숫자를 입력해주세요.")
            return

        if distance <= 0 or efficiency <= 0 or price <= 0:
            QMessageBox.warning(self, "입력 오류", "주행거리, 연비, 유가는 0보다 커야 합니다.")
            return

        total_distance = distance * 2 if self.trip_round_trip.isChecked() else distance
        fuel_amount = total_distance / efficiency
        fuel_cost = fuel_amount * price
        extra_cost = toll + parking + other
        total_cost = fuel_cost + extra_cost
        per_km = fuel_cost / total_distance
        per_100km = per_km * 100

        self.res_distance.setText(f"{total_distance:,.1f}")
        self.res_fuel.setText(f"{fuel_amount:,.2f}")
        self.res_fuel_cost.setText(f"{round(fuel_cost):,}")
        self.res_per_km.setText(f"{per_km:,.2f}")
        self.res_per_100km.setText(f"{per_100km:,.0f}")
        self.res_extra.setText(f"{round(extra_cost):,}")
        self.res_total.setText(f"{round(total_cost):,}")

    def save_to_excel(self):
        if self.res_total.text() == "-":
            QMessageBox.warning(self, "저장 불가", "먼저 '계산하기'를 실행하여 결과 값을 생성해주세요.")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "엑셀 파일 저장", "주유비_계산결과.xlsx", "Excel Files (*.xlsx)")
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "주유비 정산"

            font_title = Font(name="맑은 고딕", size=14, bold=True)
            font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
            font_data = Font(name="맑은 고딕", size=10)
            font_total = Font(name="맑은 고딕", size=12, bold=True, color="D32F2F")

            fill_header = PatternFill(start_color="163A5F", end_color="163A5F", fill_type="solid")
            align_center = Alignment(horizontal="center", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin', color='D0D7DE'),
                right=Side(style='thin', color='D0D7DE'),
                top=Side(style='thin', color='D0D7DE'),
                bottom=Side(style='thin', color='D0D7DE')
            )

            ws.merge_cells("A1:B1")
            ws["A1"] = "🚗 주유비 계산기 정산 내역"
            ws["A1"].font = font_title
            ws["A1"].alignment = align_center

            data_rows = [
                ("차량명", self.vehicle_combo.currentText() if self.vehicle_combo.currentText() else "직접 입력"),
                ("유종", self.fuel_combo.currentText()),
                ("입력 주행거리", f"{self.distance_input.text()} km"),
                ("주행 방식", "왕복" if self.trip_round_trip.isChecked() else "편도"),
                ("연비", f"{self.efficiency_input.text()} km/L"),
                ("적용 유가", f"{self.price_input.text()} 원/L"),
                ("통행료", f"{self.toll_input.text()} 원"),
                ("주차비", f"{self.parking_input.text()} 원"),
                ("기타 비용", f"{self.other_input.text()} 원"),
                ("-" * 20, "-" * 20),
                ("총 주행거리", f"{self.res_distance.text()} km"),
                ("예상 주유량", f"{self.res_fuel.text()} L"),
                ("예상 주유비", f"{self.res_fuel_cost.text()} 원"),
                ("1km당 연료비", f"{self.res_per_km.text()} 원"),
                ("추가 비용 합계", f"{self.res_extra.text()} 원"),
                ("총 여행비", f"{self.res_total.text()} 원"),
            ]

            ws.append([])
            ws.append(["항목", "내용"])

            for col_num in range(1, 3):
                cell = ws.cell(row=3, column=col_num)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border

            current_row = 4
            for item, val in data_rows:
                ws.append([item, val])
                c1 = ws.cell(row=current_row, column=1)
                c2 = ws.cell(row=current_row, column=2)
                
                c1.font = font_data
                c2.font = font_data
                c1.border = thin_border
                c2.border = thin_border
                c1.alignment = align_left
                c2.alignment = align_right

                if item == "총 여행비":
                    c1.font = font_total
                    c2.font = font_total

                current_row += 1

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 25

            wb.save(file_path)
            QMessageBox.information(self, "저장 완료", f"엑셀 파일이 성공적으로 저장되었습니다.\n경로: {file_path}")

        except Exception as e:
            QMessageBox.critical(self, "저장 오류", f"엑셀을 저장하는 중 오류가 발생했습니다.\n{e}")

    def reset_fields(self):
        self.distance_input.clear()
        self.efficiency_input.clear()
        self.trip_one_way.setChecked(True)
        self.toll_input.setText("0")
        self.parking_input.setText("0")
        self.other_input.setText("0")
        
        self.res_distance.setText("-")
        self.res_fuel.setText("-")
        self.res_fuel_cost.setText("-")
        self.res_per_km.setText("-")
        self.res_per_100km.setText("-")
        self.res_extra.setText("-")
        self.res_total.setText("-")


def main():
    app = QApplication(sys.argv)
    
    font = QGuiFont("Malgun Gothic", 10)
    app.setFont(font)

    window = FuelCalculatorWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()