import os
import sys
import time
import sqlite3
import traceback
import serial
import serial.tools.list_ports

from collections import deque
from datetime import datetime, timedelta

from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QFrame,
    QPushButton,
    QDialog,
    QMessageBox,
    QScrollArea,
    QGroupBox,
    QComboBox,
)

from PyQt5.QtCore import (
    QThread,
    pyqtSignal,
    Qt,
    QPoint,
)

from PyQt5.QtGui import QFont

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter


# ============================================================
# 기본 경로
# ============================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


# ============================================================
# 1. 설정
# ============================================================

class Config:
    # Serial
    BAUD_RATE = 9600
    SERIAL_TIMEOUT = 0.1

    # 센서 개수
    MAX_SENSORS = 4

    # UI smoothing
    # 1 = smoothing 없음
    SMOOTHING_WINDOW = 1

    # 데이터 없음 판단
    NO_DATA_TIMEOUT = 30.0

    # 재연결 간격
    RECONNECT_DELAY_MS = 3000

    # 데이터 보관 기간
    LOG_RETENTION_DAYS = 30

    # 종료 대기 시간
    THREAD_WAIT_MS = 5000

    # --------------------------------------------------------
    # 센서별 Calibration
    #
    # 최종 계산:
    # (Raw + Offset) * Scale
    # --------------------------------------------------------

    SENSOR_CALIBRATION = {
        0: {
            "scale": (1.0, 2.0, 1.0),
            "offset": (0.0, 1.0, 0.0),
        },
        1: {
            "scale": (1.0, 3.0, 1.0),
            "offset": (0.0, 1.0, 0.0),
        },
        2: {
            "scale": (1.0, 1.0, 1.0),
            "offset": (0.0, 0.0, 0.0),
        },
        3: {
            "scale": (1.0, 1.0, 1.0),
            "offset": (0.0, 0.0, 0.0),
        },
    }

    # --------------------------------------------------------
    # PM10
    # --------------------------------------------------------

    PM10_LEVELS = [
        {
            "name": "좋음",
            "min": 0,
            "max": 30,
            "color": "#28A745",
        },
        {
            "name": "보통",
            "min": 31,
            "max": 80,
            "color": "#FFD700",
        },
        {
            "name": "민감군",
            "min": 81,
            "max": 120,
            "color": "#FD7E14",
        },
        {
            "name": "나쁨",
            "min": 121,
            "max": 150,
            "color": "#DC3545",
        },
        {
            "name": "매우 나쁨",
            "min": 151,
            "max": 300,
            "color": "#800080",
        },
        {
            "name": "위험",
            "min": 301,
            "max": 600,
            "color": "#795548",
        },
    ]

    # --------------------------------------------------------
    # PM2.5
    # --------------------------------------------------------

    PM25_LEVELS = [
        {
            "name": "좋음",
            "min": 0,
            "max": 15,
            "color": "#28A745",
        },
        {
            "name": "보통",
            "min": 16,
            "max": 35,
            "color": "#FFD700",
        },
        {
            "name": "민감군",
            "min": 36,
            "max": 50,
            "color": "#FD7E14",
        },
        {
            "name": "나쁨",
            "min": 51,
            "max": 75,
            "color": "#DC3545",
        },
        {
            "name": "매우 나쁨",
            "min": 76,
            "max": 100,
            "color": "#800080",
        },
        {
            "name": "위험",
            "min": 101,
            "max": 500,
            "color": "#795548",
        },
    ]

    # --------------------------------------------------------
    # PM1.0
    # --------------------------------------------------------

    PM1_LEVELS = [
        {
            "name": "좋음",
            "min": 0,
            "max": 10,
            "color": "#28A745",
        },
        {
            "name": "보통",
            "min": 11,
            "max": 25,
            "color": "#FFD700",
        },
        {
            "name": "민감군",
            "min": 26,
            "max": 35,
            "color": "#FD7E14",
        },
        {
            "name": "나쁨",
            "min": 36,
            "max": 50,
            "color": "#DC3545",
        },
        {
            "name": "매우 나쁨",
            "min": 51,
            "max": 75,
            "color": "#800080",
        },
        {
            "name": "위험",
            "min": 76,
            "max": 300,
            "color": "#795548",
        },
    ]


# ============================================================
# 2. 데이터 Parser
# ============================================================

class DustParser:

    @staticmethod
    def parse(raw_data, calib_params):
        """
        입력 예:
            PM1,PM2.5,PM10

        예:
            10,20,30

        반환:
            {
                "pm10": 30,
                "pm25": 20,
                "pm1": 10
            }

        잘못된 데이터:
            None

        범위 초과:
            "OUT_OF_RANGE"
        """

        try:
            raw_data = raw_data.strip()

            if not raw_data:
                return None

            # 불필요한 공백만 제거
            parts = [
                item.strip()
                for item in raw_data.split(",")
            ]

            if len(parts) < 3:
                return None

            scales = calib_params.get(
                "scale",
                (1.0, 1.0, 1.0)
            )

            offsets = calib_params.get(
                "offset",
                (0.0, 0.0, 0.0)
            )

            scale_pm10, scale_pm25, scale_pm1 = scales
            offset_pm10, offset_pm25, offset_pm1 = offsets

            # 센서 데이터 순서:
            # PM1, PM2.5, PM10

            raw_pm1 = float(parts[0])
            raw_pm25 = float(parts[1])
            raw_pm10 = float(parts[2])

            pm10 = max(
                0,
                int(
                    round(
                        (raw_pm10 + offset_pm10)
                        * scale_pm10
                    )
                )
            )

            pm25 = max(
                0,
                int(
                    round(
                        (raw_pm25 + offset_pm25)
                        * scale_pm25
                    )
                )
            )

            pm1 = max(
                0,
                int(
                    round(
                        (raw_pm1 + offset_pm1)
                        * scale_pm1
                    )
                )
            )

            # 비정상 범위
            if (
                pm1 > 1000
                or pm25 > 1000
                or pm10 > 2000
            ):
                return "OUT_OF_RANGE"

            return {
                "pm10": pm10,
                "pm25": pm25,
                "pm1": pm1,
            }

        except (ValueError, TypeError, IndexError):
            return None

        except Exception:
            return None


# ============================================================
# 3. SQLite / Logging
# ============================================================

class SensorLogger:

    def __init__(self, port_name, sensor_index):

        self.port_name = port_name
        self.sensor_index = sensor_index

        self.log_dir = os.path.join(
            BASE_DIR,
            "Logs"
        )

        self.excel_dir = os.path.join(
            BASE_DIR,
            "Excel_Logs"
        )

        self.db_dir = os.path.join(
            BASE_DIR,
            "Data"
        )

        os.makedirs(
            self.log_dir,
            exist_ok=True
        )

        os.makedirs(
            self.excel_dir,
            exist_ok=True
        )

        os.makedirs(
            self.db_dir,
            exist_ok=True
        )

        self.db_path = os.path.join(
            self.db_dir,
            "dust_measurement.db"
        )

    # --------------------------------------------------------
    # Database 초기화
    # --------------------------------------------------------

    @staticmethod
    def init_global_database(db_path):

        try:

            os.makedirs(
                os.path.dirname(db_path),
                exist_ok=True
            )

            with sqlite3.connect(
                db_path,
                timeout=10.0
            ) as conn:

                conn.execute(
                    "PRAGMA journal_mode=WAL;"
                )

                conn.execute(
                    "PRAGMA synchronous=NORMAL;"
                )

                conn.execute(
                    """
                    CREATE TABLE IF NOT EXISTS measurements (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        measured_at TEXT NOT NULL,
                        sensor_index INTEGER NOT NULL,
                        port TEXT NOT NULL,
                        pm10 REAL,
                        pm25 REAL,
                        pm1 REAL,
                        status TEXT DEFAULT 'NORMAL'
                    )
                    """
                )

                # 시간 검색용
                conn.execute(
                    """
                    CREATE INDEX IF NOT EXISTS
                    idx_measurements_time
                    ON measurements(measured_at)
                    """
                )

                # 센서 검색용
                conn.execute(
                    """
                    CREATE INDEX IF NOT EXISTS
                    idx_measurements_sensor
                    ON measurements(sensor_index)
                    """
                )

                # 가장 중요한 복합 인덱스
                conn.execute(
                    """
                    CREATE INDEX IF NOT EXISTS
                    idx_measurements_sensor_time
                    ON measurements(
                        sensor_index,
                        measured_at
                    )
                    """
                )

                conn.commit()

        except Exception as e:

            print(
                f"Global SQLite 초기화 오류: {e}"
            )

    # --------------------------------------------------------
    # 오래된 로그 / DB 정리
    # --------------------------------------------------------

    @staticmethod
    def cleanup_old_logs_global():

        threshold_time = (
            time.time()
            - (
                Config.LOG_RETENTION_DAYS
                * 24
                * 60
                * 60
            )
        )

        # ----------------------------------------------------
        # 1. 파일 정리
        # ----------------------------------------------------

        target_dirs = [
            os.path.join(
                BASE_DIR,
                "Logs"
            ),
            os.path.join(
                BASE_DIR,
                "Excel_Logs"
            ),
        ]

        for target_dir in target_dirs:

            if not os.path.exists(target_dir):
                continue

            try:

                filenames = os.listdir(
                    target_dir
                )

            except Exception:
                continue

            for filename in filenames:

                file_path = os.path.join(
                    target_dir,
                    filename
                )

                if not os.path.isfile(file_path):
                    continue

                try:

                    if (
                        os.path.getmtime(file_path)
                        < threshold_time
                    ):
                        os.remove(file_path)

                except Exception:
                    pass

        # ----------------------------------------------------
        # 2. SQLite 오래된 데이터 정리
        # ----------------------------------------------------

        db_path = os.path.join(
            BASE_DIR,
            "Data",
            "dust_measurement.db"
        )

        if not os.path.exists(db_path):
            return

        try:

            threshold_datetime = (
                datetime.fromtimestamp(
                    threshold_time
                )
                .strftime(
                    "%Y-%m-%d %H:%M:%S"
                )
            )

            deleted_count = 0

            with sqlite3.connect(
                db_path,
                timeout=10.0
            ) as conn:

                cursor = conn.execute(
                    """
                    DELETE FROM measurements
                    WHERE measured_at < ?
                    """,
                    (threshold_datetime,)
                )

                deleted_count = cursor.rowcount

                conn.commit()

            # 삭제된 데이터가 있을 때만 VACUUM
            # 매번 실행하지 않도록 함
            if deleted_count > 0:

                try:

                    with sqlite3.connect(
                        db_path,
                        timeout=10.0
                    ) as conn:

                        conn.execute(
                            "VACUUM"
                        )

                except Exception as vacuum_error:

                    print(
                        f"SQLite VACUUM 오류: "
                        f"{vacuum_error}"
                    )

        except Exception as e:

            print(
                f"DB 오래된 데이터 삭제 오류: {e}"
            )

    # --------------------------------------------------------
    # Error Log
    # --------------------------------------------------------

    def write_error(self, message):

        log_file = os.path.join(
            self.log_dir,
            f"error_log_Sensor"
            f"{self.sensor_index + 1}.txt"
        )

        timestamp = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        try:

            with open(
                log_file,
                "a",
                encoding="utf-8-sig"
            ) as f:

                f.write(
                    f"{timestamp} | "
                    f"포트: {self.port_name} | "
                    f"{message}\n"
                )

        except Exception:
            pass

    # --------------------------------------------------------
    # Batch 저장
    # --------------------------------------------------------

    def save_measurements_batch(
        self,
        measurements
    ):

        if not measurements:
            return True

        try:

            with sqlite3.connect(
                self.db_path,
                timeout=10.0
            ) as conn:

                conn.execute(
                    "PRAGMA busy_timeout=10000;"
                )

                conn.executemany(
                    """
                    INSERT INTO measurements (
                        measured_at,
                        sensor_index,
                        port,
                        pm10,
                        pm25,
                        pm1,
                        status
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    measurements
                )

                conn.commit()

            return True

        except Exception as e:

            self.write_error(
                f"SQLite 일괄 저장 오류: {e}"
            )

            return False

    # --------------------------------------------------------
    # Excel Export
    # --------------------------------------------------------

    def export_excel(
        self,
        date_str=None
    ):

        try:

            if not date_str:

                date_str = (
                    datetime.now()
                    .strftime("%Y-%m-%d")
                )

            # 날짜 범위
            start_datetime = (
                f"{date_str} 00:00:00"
            )

            next_date = (
                datetime.strptime(
                    date_str,
                    "%Y-%m-%d"
                )
                + timedelta(days=1)
            ).strftime("%Y-%m-%d")

            end_datetime = (
                f"{next_date} 00:00:00"
            )

            # ------------------------------------------------
            # DB 조회
            # ------------------------------------------------

            with sqlite3.connect(
                self.db_path,
                timeout=10.0
            ) as conn:

                rows = conn.execute(
                    """
                    SELECT
                        measured_at,
                        port,
                        pm10,
                        pm25,
                        pm1,
                        status
                    FROM measurements
                    WHERE sensor_index = ?
                      AND measured_at >= ?
                      AND measured_at < ?
                    ORDER BY measured_at
                    """,
                    (
                        self.sensor_index,
                        start_datetime,
                        end_datetime,
                    )
                ).fetchall()

            if not rows:
                return False

            # ------------------------------------------------
            # Excel 생성
            # ------------------------------------------------

            filename = (
                f"Dust_log_{date_str}_"
                f"Sensor{self.sensor_index + 1}.xlsx"
            )

            file_path = os.path.join(
                self.excel_dir,
                filename
            )

            wb = Workbook()

            ws = wb.active

            ws.title = (
                f"{date_str} 측정 데이터"
            )

            headers = [
                "측정일시",
                "포트",
                "PM10",
                "PM2.5",
                "PM1.0",
                "상태",
            ]

            ws.append(headers)

            ws.row_dimensions[1].height = 25

            # ------------------------------------------------
            # 스타일
            # ------------------------------------------------

            thin_border = Border(
                left=Side(
                    style="thin",
                    color="D3D3D3"
                ),
                right=Side(
                    style="thin",
                    color="D3D3D3"
                ),
                top=Side(
                    style="thin",
                    color="D3D3D3"
                ),
                bottom=Side(
                    style="thin",
                    color="D3D3D3"
                )
            )

            header_fill = PatternFill(
                start_color="1F4E78",
                end_color="1F4E78",
                fill_type="solid"
            )

            header_font = Font(
                name="Malgun Gothic",
                size=10,
                bold=True,
                color="FFFFFF"
            )

            data_font = Font(
                name="Malgun Gothic",
                size=9
            )

            center_align = Alignment(
                horizontal="center",
                vertical="center"
            )

            right_align = Alignment(
                horizontal="right",
                vertical="center"
            )

            # ------------------------------------------------
            # Header
            # ------------------------------------------------

            for col_num in range(
                1,
                len(headers) + 1
            ):

                cell = ws.cell(
                    row=1,
                    column=col_num
                )

                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # ------------------------------------------------
            # Data
            # ------------------------------------------------

            for row_data in rows:

                ws.append(
                    list(row_data)
                )

                current_row = (
                    ws.max_row
                )

                ws.row_dimensions[
                    current_row
                ].height = 18

                for col_num, value in enumerate(
                    row_data,
                    1
                ):

                    cell = ws.cell(
                        row=current_row,
                        column=col_num
                    )

                    cell.font = data_font
                    cell.border = thin_border

                    if col_num in [1, 2, 6]:

                        cell.alignment = (
                            center_align
                        )

                    else:

                        cell.alignment = (
                            right_align
                        )

                        if isinstance(
                            value,
                            (int, float)
                        ):
                            cell.number_format = (
                                "#,##0"
                            )

            # ------------------------------------------------
            # Column Width
            # ------------------------------------------------

            for col in ws.columns:

                max_len = 0

                col_letter = (
                    get_column_letter(
                        col[0].column
                    )
                )

                for cell in col:

                    value_str = str(
                        cell.value or ""
                    )

                    if cell.row == 1:

                        length = len(
                            value_str.encode(
                                "utf-8"
                            )
                        )

                    else:

                        length = len(
                            value_str
                        )

                    max_len = max(
                        max_len,
                        length
                    )

                ws.column_dimensions[
                    col_letter
                ].width = max(
                    max_len + 4,
                    12
                )

            wb.save(file_path)

            return True

        except Exception as e:

            self.write_error(
                f"Excel Export 오류: {e}"
            )

            return False


# ============================================================
# 4. Serial Thread
# ============================================================

class SerialThread(QThread):

    # sensor_index, pm10, pm25, pm1
    data_signal = pyqtSignal(
        int,
        int,
        int,
        int
    )

    # sensor_index, message
    error_signal = pyqtSignal(
        int,
        str
    )

    def __init__(
        self,
        port_name,
        sensor_index
    ):

        super().__init__()

        self.port_name = port_name
        self.sensor_index = sensor_index

        # ----------------------------------------------------
        # Calibration
        # ----------------------------------------------------

        default_calib = {
            "scale": (
                1.0,
                1.0,
                1.0
            ),
            "offset": (
                0.0,
                0.0,
                0.0
            )
        }

        self.calib_params = (
            Config.SENSOR_CALIBRATION.get(
                sensor_index,
                default_calib
            )
        )

        # ----------------------------------------------------
        # Logger
        # ----------------------------------------------------

        self.logger = SensorLogger(
            port_name,
            sensor_index
        )

        # ----------------------------------------------------
        # UI smoothing buffer
        # ----------------------------------------------------

        window_size = max(
            1,
            Config.SMOOTHING_WINDOW
        )

        self.data_buffer = {
            "pm10": deque(
                maxlen=window_size
            ),
            "pm25": deque(
                maxlen=window_size
            ),
            "pm1": deque(
                maxlen=window_size
            ),
        }

        # ----------------------------------------------------
        # 1분 집계용
        # ----------------------------------------------------

        self.minute_sum = {
            "pm10": 0,
            "pm25": 0,
            "pm1": 0,
        }

        self.minute_count = 0

        # 현재 분
        self.current_minute_bucket = (
            self._get_current_minute()
        )

    # --------------------------------------------------------
    # 현재 분 반환
    # --------------------------------------------------------

    @staticmethod
    def _get_current_minute():

        return datetime.now().replace(
            second=0,
            microsecond=0
        )

    # --------------------------------------------------------
    # Sleep
    # --------------------------------------------------------

    def safe_sleep(self, milliseconds):

        remaining = milliseconds

        while remaining > 0:

            if self.isInterruptionRequested():
                return

            sleep_time = min(
                100,
                remaining
            )

            self.msleep(
                sleep_time
            )

            remaining -= sleep_time

    # --------------------------------------------------------
    # UI Buffer 초기화
    #
    # 분 집계 데이터는 유지
    # --------------------------------------------------------

    def clear_ui_buffers(self):

        for buffer in self.data_buffer.values():

            buffer.clear()

    # --------------------------------------------------------
    # Minute Buffer 초기화
    # --------------------------------------------------------

    def clear_minute_buffers(self):

        self.minute_sum = {
            "pm10": 0,
            "pm25": 0,
            "pm1": 0,
        }

        self.minute_count = 0

    # --------------------------------------------------------
    # 전체 Buffer 초기화
    # --------------------------------------------------------

    def clear_all_buffers(self):

        self.clear_ui_buffers()
        self.clear_minute_buffers()

    # --------------------------------------------------------
    # 데이터 하나 처리
    # --------------------------------------------------------

    def add_measurement(
        self,
        pm10,
        pm25,
        pm1
    ):

        # UI smoothing
        self.data_buffer["pm10"].append(
            pm10
        )

        self.data_buffer["pm25"].append(
            pm25
        )

        self.data_buffer["pm1"].append(
            pm1
        )

        # 1분 집계
        self.minute_sum["pm10"] += pm10
        self.minute_sum["pm25"] += pm25
        self.minute_sum["pm1"] += pm1

        # 패킷 1개 = 측정 1회
        self.minute_count += 1

    # --------------------------------------------------------
    # 현재 1분 평균 저장
    # --------------------------------------------------------

    def save_current_minute_average(
        self,
        minute_bucket
    ):

        if self.minute_count <= 0:
            return True

        avg_pm10 = int(
            round(
                self.minute_sum["pm10"]
                / self.minute_count
            )
        )

        avg_pm25 = int(
            round(
                self.minute_sum["pm25"]
                / self.minute_count
            )
        )

        avg_pm1 = int(
            round(
                self.minute_sum["pm1"]
                / self.minute_count
            )
        )

        measured_at = (
            minute_bucket.strftime(
                "%Y-%m-%d %H:%M:00"
            )
        )

        batch_data = [
            (
                measured_at,
                self.sensor_index,
                self.port_name,
                avg_pm10,
                avg_pm25,
                avg_pm1,
                "NORMAL",
            )
        ]

        # DB 저장 성공 여부 확인
        success = (
            self.logger.save_measurements_batch(
                batch_data
            )
        )

        # 저장 성공한 경우에만 집계 초기화
        if success:

            self.clear_minute_buffers()

            return True

        # 실패하면 데이터 유지
        return False

    # --------------------------------------------------------
    # 분 변경 처리
    # --------------------------------------------------------

    def check_minute_boundary(self):

        new_bucket = (
            self._get_current_minute()
        )

        if new_bucket == self.current_minute_bucket:
            return

        # 이전 분 저장
        save_success = (
            self.save_current_minute_average(
                self.current_minute_bucket
            )
        )

        # DB 저장 실패 시
        # 기존 데이터를 유지하고 현재 bucket을 변경하지 않음
        if not save_success:

            self.logger.write_error(
                "이전 분 데이터 저장 실패 - "
                "다음 루프에서 재시도"
            )

            return

        self.current_minute_bucket = (
            new_bucket
        )

    # --------------------------------------------------------
    # 현재 UI 평균
    # --------------------------------------------------------

    def get_ui_averages(self):

        try:

            pm10_buffer = (
                self.data_buffer["pm10"]
            )

            pm25_buffer = (
                self.data_buffer["pm25"]
            )

            pm1_buffer = (
                self.data_buffer["pm1"]
            )

            if not pm10_buffer:
                return None

            if not pm25_buffer:
                return None

            if not pm1_buffer:
                return None

            return (
                int(
                    sum(pm10_buffer)
                    / len(pm10_buffer)
                ),
                int(
                    sum(pm25_buffer)
                    / len(pm25_buffer)
                ),
                int(
                    sum(pm1_buffer)
                    / len(pm1_buffer)
                ),
            )

        except Exception:

            return None

    # --------------------------------------------------------
    # Thread Main
    # --------------------------------------------------------

    def run(self):

        is_connected = False

        no_data_error_sent = False

        out_of_range_error_sent = False

        last_ui_update_time = 0.0

        last_data_time = time.time()

        while not self.isInterruptionRequested():

            ser = None

            try:

                # ------------------------------------------------
                # Serial 연결
                # ------------------------------------------------

                ser = serial.Serial(
                    self.port_name,
                    Config.BAUD_RATE,
                    timeout=Config.SERIAL_TIMEOUT
                )

                ser.reset_input_buffer()

                if not is_connected:

                    self.logger.write_error(
                        f"통신 연결 성공 "
                        f"({self.port_name})"
                    )

                is_connected = True

                no_data_error_sent = False
                out_of_range_error_sent = False

                last_data_time = time.time()

                # 연결이 새로 되었을 때
                # 현재 시각을 기준으로 새 bucket 시작
                self.current_minute_bucket = (
                    self._get_current_minute()
                )

                # UI buffer만 초기화
                self.clear_ui_buffers()

                # ------------------------------------------------
                # Serial Loop
                # ------------------------------------------------

                while (
                    ser.is_open
                    and not self.isInterruptionRequested()
                ):

                    parsed_values = []

                    # --------------------------------------------
                    # Serial 데이터 읽기
                    # --------------------------------------------

                    while ser.in_waiting > 0:

                        raw_data = (
                            ser.readline()
                            .decode(
                                "utf-8",
                                errors="ignore"
                            )
                            .strip()
                        )

                        if not raw_data:
                            continue

                        parsed = (
                            DustParser.parse(
                                raw_data,
                                self.calib_params
                            )
                        )

                        # ----------------------------------------
                        # 범위 초과
                        # ----------------------------------------

                        if parsed == "OUT_OF_RANGE":

                            if not out_of_range_error_sent:

                                self.error_signal.emit(
                                    self.sensor_index,
                                    "측정값 범위 초과"
                                )

                                self.logger.write_error(
                                    f"범위 초과 Raw: "
                                    f"{raw_data}"
                                )

                                out_of_range_error_sent = True

                            continue

                        # ----------------------------------------
                        # 정상 데이터
                        # ----------------------------------------

                        if parsed is not None:

                            parsed_values.append(
                                parsed
                            )

                            last_data_time = (
                                time.time()
                            )

                            # 정상 데이터가 들어오면
                            # 범위 초과 상태 해제
                            out_of_range_error_sent = (
                                False
                            )

                    # ------------------------------------------------
                    # 중요:
                    # 데이터가 없어도 분 경계를 확인
                    # ------------------------------------------------

                    self.check_minute_boundary()

                    # ------------------------------------------------
                    # 데이터 없음 판단
                    # ------------------------------------------------

                    if (
                        time.time()
                        - last_data_time
                        >= Config.NO_DATA_TIMEOUT
                    ):

                        if not no_data_error_sent:

                            self.error_signal.emit(
                                self.sensor_index,
                                "데이터 없음"
                            )

                            self.logger.write_error(
                                f"{Config.NO_DATA_TIMEOUT:.0f}초 "
                                f"이상 데이터 없음"
                            )

                            no_data_error_sent = True

                            # UI만 초기화
                            # 현재 분 집계 데이터는 유지
                            self.clear_ui_buffers()

                    # ------------------------------------------------
                    # 정상 데이터 처리
                    # ------------------------------------------------

                    elif parsed_values:

                        no_data_error_sent = False

                        for parsed in parsed_values:

                            self.add_measurement(
                                parsed["pm10"],
                                parsed["pm25"],
                                parsed["pm1"]
                            )

                        # --------------------------------------------
                        # UI 업데이트는 최대 1초에 한 번
                        # --------------------------------------------

                        current_time = (
                            time.time()
                        )

                        if (
                            current_time
                            - last_ui_update_time
                            >= 1.0
                        ):

                            last_ui_update_time = (
                                current_time
                            )

                            averages = (
                                self.get_ui_averages()
                            )

                            if averages is not None:

                                self.data_signal.emit(
                                    self.sensor_index,
                                    averages[0],
                                    averages[1],
                                    averages[2],
                                )

                    # ------------------------------------------------
                    # CPU 점유율 방지
                    # ------------------------------------------------

                    self.safe_sleep(100)

            # ====================================================
            # Serial 오류
            # ====================================================

            except serial.SerialException as e:

                if is_connected:

                    self.logger.write_error(
                        f"시리얼 통신 오류: {e}"
                    )

                is_connected = False

                # UI buffer만 초기화
                # 분 집계 데이터는 유지
                self.clear_ui_buffers()

                self.error_signal.emit(
                    self.sensor_index,
                    "연결 실패/끊김"
                )

            # ====================================================
            # 기타 오류
            # ====================================================

            except Exception:

                if is_connected:

                    self.logger.write_error(
                        "시스템 오류:\n"
                        + traceback.format_exc()
                    )

                is_connected = False

                self.clear_ui_buffers()

                self.error_signal.emit(
                    self.sensor_index,
                    "시스템 오류"
                )

            # ====================================================
            # Serial 종료
            # ====================================================

            finally:

                if ser is not None:

                    try:

                        if ser.is_open:
                            ser.close()

                    except Exception:
                        pass

                # 종료 요청이 없다면 재연결
                if not self.isInterruptionRequested():

                    self.safe_sleep(
                        Config.RECONNECT_DELAY_MS
                    )

        # ========================================================
        # Thread 종료 직전
        # 현재 분 데이터 저장
        # ========================================================

        try:

            self.save_current_minute_average(
                self.current_minute_bucket
            )

        except Exception as e:

            self.logger.write_error(
                f"Thread 종료 시 "
                f"마지막 데이터 저장 오류: {e}"
            )


# ============================================================
# 5. 미세먼지 Level Widget
# ============================================================

class DustLevelWidget(QWidget):

    def __init__(
        self,
        title="미세먼지",
        levels=None,
        parent=None
    ):

        super().__init__(parent)

        self.title = title

        self.levels = (
            levels
            if levels
            else Config.PM10_LEVELS
        )

        self.current_value = 0

        self.initUI()

    # --------------------------------------------------------
    # UI
    # --------------------------------------------------------

    def initUI(self):

        self.setStyleSheet(
            """
            background-color: transparent;
            border: none;
            """
        )

        main_layout = QVBoxLayout(self)

        main_layout.setContentsMargins(
            10,
            10,
            10,
            10
        )

        main_layout.setSpacing(3)

        # ----------------------------------------------------
        # Title
        # ----------------------------------------------------

        self.title_label = QLabel(
            self.title
        )

        self.title_label.setFont(
            QFont(
                "Malgun Gothic",
                11,
                QFont.Bold
            )
        )

        self.title_label.setAlignment(
            Qt.AlignCenter
        )

        self.title_label.setStyleSheet(
            "color: black; border: none;"
        )

        main_layout.addWidget(
            self.title_label
        )

        # ----------------------------------------------------
        # Value
        # ----------------------------------------------------

        value_layout = QHBoxLayout()

        value_layout.setAlignment(
            Qt.AlignCenter
        )

        self.val_label = QLabel(
            "----"
        )

        self.val_label.setFont(
            QFont(
                "Arial",
                28,
                QFont.Bold
            )
        )

        self.val_label.setStyleSheet(
            "color: black; border: none;"
        )

        value_layout.addWidget(
            self.val_label
        )

        unit_label = QLabel(
            "μg/m³"
        )

        unit_label.setFont(
            QFont(
                "Arial",
                10,
                QFont.Bold
            )
        )

        unit_label.setStyleSheet(
            """
            color: #666666;
            margin-bottom: 4px;
            border: none;
            """
        )

        unit_label.setAlignment(
            Qt.AlignBottom
        )

        value_layout.addWidget(
            unit_label
        )

        main_layout.addLayout(
            value_layout
        )

        # ----------------------------------------------------
        # Arrow
        # ----------------------------------------------------

        self.arrow_container = QWidget()

        self.arrow_container.setStyleSheet(
            "border: none;"
        )

        self.arrow_container.setFixedHeight(
            14
        )

        self.arrow_label = QLabel(
            "▼"
        )

        self.arrow_label.setFont(
            QFont(
                "Arial",
                9,
                QFont.Bold
            )
        )

        self.arrow_label.setStyleSheet(
            "color: black; border: none;"
        )

        self.arrow_label.setAlignment(
            Qt.AlignCenter
        )

        self.arrow_label.setFixedWidth(
            14
        )

        self.arrow_label.setParent(
            self.arrow_container
        )

        main_layout.addWidget(
            self.arrow_container
        )

        # ----------------------------------------------------
        # Level Bar
        # ----------------------------------------------------

        level_bar_frame = QFrame()

        level_bar_frame.setStyleSheet(
            "border: none;"
        )

        self.level_bar_layout = (
            QHBoxLayout(
                level_bar_frame
            )
        )

        self.level_bar_layout.setContentsMargins(
            0,
            0,
            0,
            0
        )

        self.level_bar_layout.setSpacing(
            2
        )

        self.level_bars = []

        for i, level in enumerate(
            self.levels
        ):

            bar = QFrame()

            bar.setFixedHeight(
                10
            )

            border_radius = ""

            if i == 0:

                border_radius = (
                    "border-top-left-radius: 5px;"
                    "border-bottom-left-radius: 5px;"
                )

            elif i == len(self.levels) - 1:

                border_radius = (
                    "border-top-right-radius: 5px;"
                    "border-bottom-right-radius: 5px;"
                )

            bar.setStyleSheet(
                f"""
                background-color:
                    {level['color']};
                {border_radius}
                """
            )

            self.level_bars.append(
                bar
            )

            self.level_bar_layout.addWidget(
                bar
            )

        main_layout.addWidget(
            level_bar_frame
        )

        # ----------------------------------------------------
        # Status
        # ----------------------------------------------------

        self.status_text_label = QLabel(
            "대기 중..."
        )

        self.status_text_label.setFont(
            QFont(
                "Malgun Gothic",
                11,
                QFont.Bold
            )
        )

        self.status_text_label.setAlignment(
            Qt.AlignCenter
        )

        self.status_text_label.setFixedHeight(
            30
        )

        self.status_text_label.setStyleSheet(
            """
            background-color: #E0E0E0;
            border-radius: 6px;
            color: gray;
            """
        )

        main_layout.addWidget(
            self.status_text_label
        )

    # --------------------------------------------------------
    # Value update
    # --------------------------------------------------------

    def update_val(self, value):

        self.arrow_label.show()

        self.current_value = value

        self.val_label.setText(
            str(value)
        )

        current_level = self.levels[
            self._find_level_index(value)
        ]

        self.status_text_label.setText(
            current_level["name"]
        )

        text_color = (
            "black"
            if current_level["name"] == "보통"
            else "white"
        )

        self.status_text_label.setStyleSheet(
            f"""
            background-color:
                {current_level['color']};
            border-radius: 6px;
            color: {text_color};
            border: none;
            """
        )

        self.update_arrow_position()

    # --------------------------------------------------------
    # Level index
    # --------------------------------------------------------

    def _find_level_index(self, value):

        for i, level in enumerate(
            self.levels
        ):

            if value <= level["max"]:
                return i

        return len(self.levels) - 1

    # --------------------------------------------------------
    # Error
    # --------------------------------------------------------

    def set_error_state(self, msg):

        self.val_label.setText(
            "----"
        )

        self.status_text_label.setText(
            msg
        )

        self.status_text_label.setStyleSheet(
            """
            background-color: #FFCDD2;
            border-radius: 6px;
            color: #B71C1C;
            border: none;
            """
        )

        self.current_value = (
            self.levels[0]["min"]
        )

        self.update_arrow_position()

        self.arrow_label.hide()

    # --------------------------------------------------------
    # Arrow
    # --------------------------------------------------------

    def update_arrow_position(self):

        if not self.level_bars:
            return

        if (
            self.level_bars[0]
            .geometry()
            .width()
            == 0
        ):
            return

        value = self.current_value

        level_index = (
            self._find_level_index(
                value
            )
        )

        current_level = (
            self.levels[level_index]
        )

        level_min = (
            current_level["min"]
        )

        level_max = (
            current_level["max"]
        )

        if value <= level_min:

            ratio = 0.0

        elif value >= level_max:

            ratio = 1.0

        else:

            denominator = (
                level_max - level_min
            )

            if denominator <= 0:

                ratio = 0.0

            else:

                ratio = (
                    (value - level_min)
                    / denominator
                )

        target_bar = (
            self.level_bars[
                level_index
            ]
        )

        target_x = (
            target_bar.geometry().x()
            + (
                target_bar.geometry().width()
                * ratio
            )
        )

        arrow_x = int(
            target_x
            - (
                self.arrow_label.width()
                / 2
            )
        )

        max_x = (
            self.arrow_container.width()
            - self.arrow_label.width()
        )

        arrow_x = max(
            0,
            min(
                arrow_x,
                max_x
            )
        )

        self.arrow_label.move(
            QPoint(
                arrow_x,
                int(
                    self.arrow_container.height()
                    - self.arrow_label.height()
                    + 2
                )
            )
        )

    # --------------------------------------------------------
    # Resize
    # --------------------------------------------------------

    def resizeEvent(self, event):

        super().resizeEvent(event)

        self.update_arrow_position()


# ============================================================
# 6. Port Selection Dialog
# ============================================================

class PortSelectionDialog(QDialog):

    def __init__(
        self,
        parent=None
    ):

        super().__init__(
            parent
        )

        self.selected_ports = {}

        self.available_ports = sorted(
            [
                port.device
                for port
                in serial.tools.list_ports.comports()
            ]
        )

        self.initUI()

    # --------------------------------------------------------
    # UI
    # --------------------------------------------------------

    def initUI(self):

        self.setWindowTitle(
            "센서 포트 설정"
        )

        self.resize(
            350,
            250
        )

        self.setStyleSheet(
            "background-color: white;"
        )

        layout = QVBoxLayout(
            self
        )

        layout.setContentsMargins(
            20,
            20,
            20,
            20
        )

        layout.setSpacing(
            15
        )

        title_label = QLabel(
            "모니터링할 센서 포트를 선택하세요"
        )

        title_label.setFont(
            QFont(
                "Malgun Gothic",
                10,
                QFont.Bold
            )
        )

        layout.addWidget(
            title_label
        )

        # ----------------------------------------------------
        # Sensor Combo
        # ----------------------------------------------------

        self.combos = []

        for i in range(
            Config.MAX_SENSORS
        ):

            h_layout = QHBoxLayout()

            label = QLabel(
                f"센서 {i + 1} 포트:"
            )

            label.setFont(
                QFont(
                    "Malgun Gothic",
                    9
                )
            )

            combo = QComboBox()

            combo.addItem(
                "선택 안 함"
            )

            combo.addItems(
                self.available_ports
            )

            combo.setFont(
                QFont(
                    "Malgun Gothic",
                    9
                )
            )

            combo.currentIndexChanged.connect(
                self.update_combo_items
            )

            h_layout.addWidget(
                label
            )

            h_layout.addWidget(
                combo
            )

            layout.addLayout(
                h_layout
            )

            self.combos.append(
                combo
            )

        # ----------------------------------------------------
        # Start button
        # ----------------------------------------------------

        self.start_btn = QPushButton(
            "모니터링 시작"
        )

        self.start_btn.setFixedHeight(
            35
        )

        self.start_btn.setFont(
            QFont(
                "Malgun Gothic",
                10,
                QFont.Bold
            )
        )

        self.start_btn.setStyleSheet(
            """
            background-color: #007BFF;
            color: white;
            border-radius: 4px;
            """
        )

        self.start_btn.clicked.connect(
            self.accept_selection
        )

        layout.addWidget(
            self.start_btn
        )

    # --------------------------------------------------------
    # 중복 포트 방지
    # --------------------------------------------------------

    def update_combo_items(self):

        selected_values = [
            combo.currentText()
            for combo in self.combos
            if (
                combo.currentText()
                and combo.currentText()
                != "선택 안 함"
            )
        ]

        for combo in self.combos:

            current_selection = (
                combo.currentText()
            )

            combo.blockSignals(
                True
            )

            combo.clear()

            combo.addItem(
                "선택 안 함"
            )

            for port in self.available_ports:

                if (
                    port in selected_values
                    and port
                    != current_selection
                ):
                    continue

                combo.addItem(
                    port
                )

            index = combo.findText(
                current_selection
            )

            if index != -1:

                combo.setCurrentIndex(
                    index
                )

            else:

                combo.setCurrentIndex(
                    0
                )

            combo.blockSignals(
                False
            )

    # --------------------------------------------------------
    # Selection
    # --------------------------------------------------------

    def accept_selection(self):

        self.selected_ports = {
            i: combo.currentText()
            for i, combo in enumerate(
                self.combos
            )
            if (
                combo.currentText()
                and combo.currentText()
                != "선택 안 함"
            )
        }

        if not self.selected_ports:

            QMessageBox.warning(
                self,
                "경고",
                "최소 하나 이상의 센서 포트를 선택해 주세요."
            )

            return

        self.accept()


# ============================================================
# 7. Main Monitoring Window
# ============================================================

class DustMonitorApp(QMainWindow):

    def __init__(
        self,
        slot_mapping
    ):

        super().__init__()

        self.threads = []

        self.sensor_widgets = {}

        self.slot_mapping = slot_mapping

        self.initUI()

        self.start_monitoring()

    # --------------------------------------------------------
    # UI
    # --------------------------------------------------------

    def initUI(self):

        self.setWindowTitle(
            "다중 미세먼지 모니터링 시스템"
        )

        self.setStyleSheet(
            """
            QMainWindow {
                background-color: white;
            }
            """
        )

        self.setMinimumSize(
            400,
            150
        )

        scroll = QScrollArea()

        scroll.setWidgetResizable(
            True
        )

        self.setCentralWidget(
            scroll
        )

        scroll_content = QWidget()

        scroll.setWidget(
            scroll_content
        )

        self.scroll_layout = (
            QVBoxLayout(
                scroll_content
            )
        )

        self.scroll_layout.setContentsMargins(
            15,
            15,
            15,
            15
        )

        self.scroll_layout.setSpacing(
            15
        )

        self.scroll_layout.addStretch(
            1
        )

    # --------------------------------------------------------
    # Monitoring Start
    # --------------------------------------------------------

    def start_monitoring(self):

        for sensor_index, port_name in (
            self.slot_mapping.items()
        ):

            # ------------------------------------------------
            # Group
            # ------------------------------------------------

            group_box = QGroupBox(
                f"센서 {sensor_index + 1} "
                f"포트: {port_name}"
            )

            group_box.setFont(
                QFont(
                    "Malgun Gothic",
                    11,
                    QFont.Bold
                )
            )

            port_layout = QHBoxLayout(
                group_box
            )

            port_layout.setSpacing(
                10
            )

            # ------------------------------------------------
            # PM10
            # ------------------------------------------------

            w_pm10 = DustLevelWidget(
                "PM10 (미세먼지)",
                Config.PM10_LEVELS
            )

            # ------------------------------------------------
            # PM2.5
            # ------------------------------------------------

            w_pm25 = DustLevelWidget(
                "PM2.5 (초미세먼지)",
                Config.PM25_LEVELS
            )

            # ------------------------------------------------
            # PM1
            # ------------------------------------------------

            w_pm1 = DustLevelWidget(
                "PM1.0 (극미세먼지)",
                Config.PM1_LEVELS
            )

            port_layout.addWidget(
                w_pm10
            )

            port_layout.addWidget(
                w_pm25
            )

            port_layout.addWidget(
                w_pm1
            )

            # ------------------------------------------------
            # Layout에 추가
            # ------------------------------------------------

            self.scroll_layout.insertWidget(
                self.scroll_layout.count() - 1,
                group_box
            )

            self.sensor_widgets[
                sensor_index
            ] = {
                "pm10": w_pm10,
                "pm25": w_pm25,
                "pm1": w_pm1,
            }

            # ------------------------------------------------
            # Serial Thread
            # ------------------------------------------------

            thread = SerialThread(
                port_name,
                sensor_index
            )

            thread.data_signal.connect(
                self.update_data
            )

            thread.error_signal.connect(
                self.handle_error
            )

            thread.start()

            self.threads.append(
                thread
            )

        # ----------------------------------------------------
        # Window Size
        # ----------------------------------------------------

        sensor_count = len(
            self.slot_mapping
        )

        calculated_height = max(
            220,
            sensor_count * 245 + 60
        )

        self.resize(
            980,
            calculated_height
        )

    # --------------------------------------------------------
    # Data Update
    # --------------------------------------------------------

    def update_data(
        self,
        sensor_index,
        pm10,
        pm25,
        pm1
    ):

        widgets = (
            self.sensor_widgets.get(
                sensor_index
            )
        )

        if not widgets:
            return

        widgets["pm10"].update_val(
            pm10
        )

        widgets["pm25"].update_val(
            pm25
        )

        widgets["pm1"].update_val(
            pm1
        )

    # --------------------------------------------------------
    # Error
    # --------------------------------------------------------

    def handle_error(
        self,
        sensor_index,
        error_msg
    ):

        widgets = (
            self.sensor_widgets.get(
                sensor_index
            )
        )

        if not widgets:
            return

        widgets["pm10"].set_error_state(
            error_msg
        )

        widgets["pm25"].set_error_state(
            error_msg
        )

        widgets["pm1"].set_error_state(
            error_msg
        )

    # --------------------------------------------------------
    # Close
    # --------------------------------------------------------

    def closeEvent(self, event):

        # ----------------------------------------------------
        # 1. 모든 Thread 종료 요청
        # ----------------------------------------------------

        for thread in self.threads:

            if thread.isRunning():

                thread.requestInterruption()

        # ----------------------------------------------------
        # 2. Thread 종료 대기
        # ----------------------------------------------------

        for thread in self.threads:

            if thread.isRunning():

                finished = thread.wait(
                    Config.THREAD_WAIT_MS
                )

                if not finished:

                    thread.logger.write_error(
                        "프로그램 종료 시 "
                        "Thread 종료 시간 초과"
                    )

        # ----------------------------------------------------
        # 3. Thread가 종료되지 않은 경우
        #    마지막 데이터 저장 시도
        # ----------------------------------------------------

        for thread in self.threads:

            try:

                if thread.isRunning():
                    continue

                # Thread 종료 과정에서 이미
                # 마지막 분 저장을 시도했으므로
                # 남은 데이터가 있으면 한 번 더 확인
                thread.save_current_minute_average(
                    thread.current_minute_bucket
                )

            except Exception as e:

                thread.logger.write_error(
                    f"종료 데이터 저장 오류: {e}"
                )

        # ----------------------------------------------------
        # 4. Excel Export
        # ----------------------------------------------------

        for thread in self.threads:

            try:

                thread.logger.export_excel()

            except Exception as e:

                thread.logger.write_error(
                    f"종료 시 Excel Export 오류: {e}"
                )

        event.accept()

    # --------------------------------------------------------
    # Window Drag
    # --------------------------------------------------------

    def mousePressEvent(
        self,
        event
    ):

        if event.button() == Qt.LeftButton:

            self.dragPos = (
                event.globalPos()
            )

    def mouseMoveEvent(
        self,
        event
    ):

        if (
            event.buttons()
            == Qt.LeftButton
            and hasattr(
                self,
                "dragPos"
            )
        ):

            self.move(
                self.pos()
                + event.globalPos()
                - self.dragPos
            )

            self.dragPos = (
                event.globalPos()
            )

    # --------------------------------------------------------
    # ESC
    # --------------------------------------------------------

    def keyPressEvent(
        self,
        event
    ):

        if event.key() == Qt.Key_Escape:

            self.close()

        else:

            super().keyPressEvent(
                event
            )


# ============================================================
# 8. Main
# ============================================================

def main():

    app = QApplication(
        sys.argv
    )

    # --------------------------------------------------------
    # DB 초기화
    # --------------------------------------------------------

    db_path = os.path.join(
        BASE_DIR,
        "Data",
        "dust_measurement.db"
    )

    SensorLogger.init_global_database(
        db_path
    )

    # --------------------------------------------------------
    # 오래된 데이터 정리
    # --------------------------------------------------------

    SensorLogger.cleanup_old_logs_global()

    # --------------------------------------------------------
    # Port Selection
    # --------------------------------------------------------

    dialog = PortSelectionDialog()

    if dialog.exec_() != QDialog.Accepted:

        return 0

    # --------------------------------------------------------
    # Main Window
    # --------------------------------------------------------

    main_window = DustMonitorApp(
        dialog.selected_ports
    )

    main_window.show()

    return app.exec_()


# ============================================================
# Entry Point
# ============================================================

if __name__ == "__main__":

    sys.exit(
        main()
    )
