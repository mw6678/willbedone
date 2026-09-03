import os
import sys
import time
import sqlite3
import traceback
import queue
import threading
import serial
import serial.tools.list_ports

from collections import deque
from datetime import datetime, timedelta

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QFrame, QPushButton, QDialog, QMessageBox, QScrollArea,
    QGroupBox, QComboBox,
)
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QPoint
from PyQt5.QtGui import QFont

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


BASE_DIR = os.path.dirname(os.path.abspath(__file__))


class Config:
    BAUD_RATE = 9600
    SERIAL_TIMEOUT = 0.1
    MAX_SENSORS = 4
    SMOOTHING_WINDOW = 1
    NO_DATA_TIMEOUT = 30.0
    RECONNECT_DELAY_MS = 3000
    LOG_RETENTION_DAYS = 30
    THREAD_WAIT_MS = 5000

    SENSOR_CALIBRATION = {
        0: {"scale": (1.0, 2.0, 1.0), "offset": (0.0, 1.0, 0.0)},
        1: {"scale": (1.0, 3.0, 1.0), "offset": (0.0, 1.0, 0.0)},
        2: {"scale": (1.0, 1.0, 1.0), "offset": (0.0, 0.0, 0.0)},
        3: {"scale": (1.0, 1.0, 1.0), "offset": (0.0, 0.0, 0.0)},
    }

    PM10_LEVELS = [
        {"name": "좋음", "min": 0, "max": 30, "color": "#28A745"},
        {"name": "보통", "min": 31, "max": 80, "color": "#FFD700"},
        {"name": "민감군", "min": 81, "max": 120, "color": "#FD7E14"},
        {"name": "나쁨", "min": 121, "max": 150, "color": "#DC3545"},
        {"name": "매우 나쁨", "min": 151, "max": 300, "color": "#800080"},
        {"name": "위험", "min": 301, "max": 600, "color": "#795548"},
    ]

    PM25_LEVELS = [
        {"name": "좋음", "min": 0, "max": 15, "color": "#28A745"},
        {"name": "보통", "min": 16, "max": 35, "color": "#FFD700"},
        {"name": "민감군", "min": 36, "max": 50, "color": "#FD7E14"},
        {"name": "나쁨", "min": 51, "max": 75, "color": "#DC3545"},
        {"name": "매우 나쁨", "min": 76, "max": 100, "color": "#800080"},
        {"name": "위험", "min": 101, "max": 500, "color": "#795548"},
    ]

    PM1_LEVELS = [
        {"name": "좋음", "min": 0, "max": 10, "color": "#28A745"},
        {"name": "보통", "min": 11, "max": 25, "color": "#FFD700"},
        {"name": "민감군", "min": 26, "max": 35, "color": "#FD7E14"},
        {"name": "나쁨", "min": 36, "max": 50, "color": "#DC3545"},
        {"name": "매우 나쁨", "min": 51, "max": 75, "color": "#800080"},
        {"name": "위험", "min": 76, "max": 300, "color": "#795548"},
    ]


class DustParser:
    @staticmethod
    def parse(raw_data, calib_params):
        try:
            raw_data = raw_data.strip()
            if not raw_data:
                return None

            parts = [item.strip() for item in raw_data.split(",")]
            if len(parts) < 3:
                return None

            scales = calib_params.get("scale", (1.0, 1.0, 1.0))
            offsets = calib_params.get("offset", (0.0, 0.0, 0.0))
            scale_pm10, scale_pm25, scale_pm1 = scales
            offset_pm10, offset_pm25, offset_pm1 = offsets

            raw_pm1 = float(parts[0])
            raw_pm25 = float(parts[1])
            raw_pm10 = float(parts[2])

            pm10 = max(0, int(round((raw_pm10 + offset_pm10) * scale_pm10)))
            pm25 = max(0, int(round((raw_pm25 + offset_pm25) * scale_pm25)))
            pm1 = max(0, int(round((raw_pm1 + offset_pm1) * scale_pm1)))

            if pm1 > 1000 or pm25 > 1000 or pm10 > 2000:
                return "OUT_OF_RANGE"

            return {"pm10": pm10, "pm25": pm25, "pm1": pm1}
        except (ValueError, TypeError, IndexError):
            return None


class DatabaseManager:
    _instance = None
    _lock = threading.Lock()

    def __new__(cls, db_path=None):
        with cls._lock:
            if cls._instance is None:
                cls._instance = super(DatabaseManager, cls).__new__(cls)
                cls._instance._initialized = False
        return cls._instance

    def __init__(self, db_path=None):
        if self._initialized:
            return
        with self._lock:
            if self._initialized:
                return
            self.db_path = db_path or os.path.join(BASE_DIR, "Data", "dust_measurement.db")
            self.queue = queue.Queue(maxsize=2000)
            self.running = True
            self._init_db()
            self.worker_thread = threading.Thread(target=self._db_worker, daemon=True)
            self.worker_thread.start()
            self._initialized = True

    def _init_db(self):
        try:
            os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
            with sqlite3.connect(self.db_path, timeout=15.0) as conn:
                conn.execute("PRAGMA journal_mode=WAL;")
                conn.execute("PRAGMA synchronous=NORMAL;")
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS measurements (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        measured_at TEXT NOT NULL,
                        sensor_index INTEGER NOT NULL,
                        port TEXT NOT NULL,
                        pm10 REAL,
                        pm25 REAL,
                        pm1 REAL,
                        status TEXT DEFAULT 'NORMAL'
                    )
                """)
                conn.execute("CREATE INDEX IF NOT EXISTS idx_measurements_time ON measurements(measured_at)")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_measurements_sensor ON measurements(sensor_index)")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_measurements_sensor_time ON measurements(sensor_index, measured_at)")
                conn.commit()
        except Exception as e:
            print(f"Global SQLite 초기화 오류: {e}")

    def _db_worker(self):
        while self.running:
            try:
                task = self.queue.get(timeout=0.5)
                if task is None:
                    self.queue.task_done()
                    break
                func, args, future_event, result_holder = task
                try:
                    with sqlite3.connect(self.db_path, timeout=15.0) as conn:
                        conn.execute("PRAGMA busy_timeout=10000;")
                        res = func(conn, *args)
                        conn.commit()
                        if result_holder is not None:
                            result_holder['result'] = res
                except Exception as e:
                    print(f"DB Worker 작업 수행 오류: {e}")
                    if result_holder is not None:
                        result_holder['error'] = e
                finally:
                    if future_event:
                        future_event.set()
                    self.queue.task_done()
            except queue.Empty:
                continue
            except Exception as e:
                print(f"DB Worker 루프 오류: {e}")

    def execute_sync(self, func, *args):
        future_event = threading.Event()
        result_holder = {}
        try:
            # 큐가 가득 찼을 때 블로킹 타임아웃을 주어 프로그램 멈춤 방지
            self.queue.put((func, args, future_event, result_holder), timeout=2.0)
        except queue.Full:
            print("DB 작업 큐가 가득 찼습니다. 일부 로그 저장이 지연될 수 있습니다.")
            return False

        future_event.wait(timeout=Config.THREAD_WAIT_MS / 2000.0)
        if 'error' in result_holder:
            return False
        return result_holder.get('result', True)

    def stop(self):
        self.running = False
        self.queue.put(None)
        self.worker_thread.join(timeout=2.0)


class SensorLogger:
    def __init__(self, port_name, sensor_index):
        self.port_name = port_name
        self.sensor_index = sensor_index
        self.log_dir = os.path.join(BASE_DIR, "Logs")
        self.excel_dir = os.path.join(BASE_DIR, "Excel_Logs")
        self.db_dir = os.path.join(BASE_DIR, "Data")

        os.makedirs(self.log_dir, exist_ok=True)
        os.makedirs(self.excel_dir, exist_ok=True)
        os.makedirs(self.db_dir, exist_ok=True)

        self.db_path = os.path.join(self.db_dir, "dust_measurement.db")
        self.db_manager = DatabaseManager(self.db_path)

    @staticmethod
    def init_global_database(db_path):
        DatabaseManager(db_path)

    @staticmethod
    def cleanup_old_logs_global():
        threshold_time = time.time() - (Config.LOG_RETENTION_DAYS * 24 * 60 * 60)
        target_dirs = [os.path.join(BASE_DIR, "Logs"), os.path.join(BASE_DIR, "Excel_Logs")]

        for target_dir in target_dirs:
            if not os.path.exists(target_dir):
                continue
            try:
                filenames = os.listdir(target_dir)
            except Exception:
                continue

            for filename in filenames:
                file_path = os.path.join(target_dir, filename)
                if not os.path.isfile(file_path):
                    continue
                try:
                    if os.path.getmtime(file_path) < threshold_time:
                        os.remove(file_path)
                except OSError:
                    pass

        db_path = os.path.join(BASE_DIR, "Data", "dust_measurement.db")
        if not os.path.exists(db_path):
            return

        try:
            threshold_datetime = datetime.fromtimestamp(threshold_time).strftime("%Y-%m-%d %H:%M:%S")

            def _delete_old(conn):
                cursor = conn.execute("DELETE FROM measurements WHERE measured_at < ?", (threshold_datetime,))
                return cursor.rowcount

            db_mgr = DatabaseManager(db_path)
            deleted_count = db_mgr.execute_sync(_delete_old)

            if isinstance(deleted_count, int) and deleted_count > 0:
                def _vacuum(conn):
                    conn.execute("VACUUM")
                db_mgr.execute_sync(_vacuum)
        except Exception as e:
            print(f"DB 오래된 데이터 삭제 오류: {e}")

    def write_error(self, message):
        log_file = os.path.join(self.log_dir, f"error_log_Sensor{self.sensor_index + 1}.txt")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with open(log_file, "a", encoding="utf-8-sig") as f:
                f.write(f"{timestamp} | 포트: {self.port_name} | {message}\n")
        except OSError:
            pass

    def save_measurements_batch(self, measurements):
        if not measurements:
            return True

        def _insert(conn, data):
            conn.executemany("""
                INSERT INTO measurements (
                    measured_at, sensor_index, port, pm10, pm25, pm1, status
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, data)
            return True

        success = self.db_manager.execute_sync(_insert, measurements)
        if not success:
            self.write_error("SQLite 일괄 저장 오류")
            return False
        return True

    def export_excel(self, date_str=None):
        try:
            if not date_str:
                date_str = datetime.now().strftime("%Y-%m-%d")

            start_datetime = f"{date_str} 00:00:00"
            next_date = (datetime.strptime(date_str, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
            end_datetime = f"{next_date} 00:00:00"

            def _fetch_chunks(conn):
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT measured_at, port, pm10, pm25, pm1, status
                    FROM measurements
                    WHERE sensor_index = ? AND measured_at >= ? AND measured_at < ?
                    ORDER BY measured_at
                """, (self.sensor_index, start_datetime, end_datetime))
                while True:
                    rows = cursor.fetchmany(1000)
                    if not rows:
                        break
                    yield rows

            filename = f"Dust_log_{date_str}_Sensor{self.sensor_index + 1}.xlsx"
            file_path = os.path.join(self.excel_dir, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = f"{date_str} 측정 데이터"

            headers = ["측정일시", "포트", "PM10", "PM2.5", "PM1.0", "상태"]
            ws.append(headers)
            ws.row_dimensions[1].height = 25

            thin_border = Border(
                left=Side(style="thin", color="D3D3D3"), right=Side(style="thin", color="D3D3D3"),
                top=Side(style="thin", color="D3D3D3"), bottom=Side(style="thin", color="D3D3D3")
            )
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
            data_font = Font(name="Malgun Gothic", size=9)
            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")

            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            has_data = False
            with sqlite3.connect(self.db_path, timeout=10.0) as conn:
                for chunk in _fetch_chunks(conn):
                    has_data = True
                    for row_data in chunk:
                        ws.append(list(row_data))
                        current_row = ws.max_row
                        ws.row_dimensions[current_row].height = 18

                        for col_num, value in enumerate(row_data, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.font = data_font
                            cell.border = thin_border
                            if col_num in [1, 2, 6]:
                                cell.alignment = center_align
                            else:
                                cell.alignment = right_align
                                if isinstance(value, (int, float)):
                                    cell.number_format = "#,##0"

            if not has_data:
                return False

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    value_str = str(cell.value or "")
                    length = len(value_str.encode("utf-8")) if cell.row == 1 else len(value_str)
                    max_len = max(max_len, length)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            wb.save(file_path)
            return True
        except Exception as e:
            self.write_error(f"Excel Export 오류: {e}")
            return False


class SerialThread(QThread):
    data_signal = pyqtSignal(int, int, int, int)
    error_signal = pyqtSignal(int, str)

    def __init__(self, port_name, sensor_index):
        super().__init__()
        self.port_name = port_name
        self.sensor_index = sensor_index

        default_calib = {"scale": (1.0, 1.0, 1.0), "offset": (0.0, 0.0, 0.0)}
        self.calib_params = Config.SENSOR_CALIBRATION.get(sensor_index, default_calib)
        self.logger = SensorLogger(port_name, sensor_index)

        window_size = max(1, Config.SMOOTHING_WINDOW)
        self.data_buffer = {
            "pm10": deque(maxlen=window_size),
            "pm25": deque(maxlen=window_size),
            "pm1": deque(maxlen=window_size),
        }

        self.minute_sum = {"pm10": 0, "pm25": 0, "pm1": 0}
        self.minute_count = 0
        self.current_minute_bucket = self._get_current_minute()

    @staticmethod
    def _get_current_minute():
        return datetime.now().replace(second=0, microsecond=0)

    def safe_sleep(self, milliseconds):
        remaining = milliseconds
        while remaining > 0:
            if self.isInterruptionRequested():
                return
            sleep_time = min(50, remaining)
            self.msleep(sleep_time)
            remaining -= sleep_time

    def clear_ui_buffers(self):
        for buffer in self.data_buffer.values():
            buffer.clear()

    def clear_minute_buffers(self):
        self.minute_sum["pm10"] = 0
        self.minute_sum["pm25"] = 0
        self.minute_sum["pm1"] = 0
        self.minute_count = 0

    def clear_all_buffers(self):
        self.clear_ui_buffers()
        self.clear_minute_buffers()

    def add_measurement(self, pm10, pm25, pm1):
        self.data_buffer["pm10"].append(pm10)
        self.data_buffer["pm25"].append(pm25)
        self.data_buffer["pm1"].append(pm1)

        self.minute_sum["pm10"] += pm10
        self.minute_sum["pm25"] += pm25
        self.minute_sum["pm1"] += pm1
        self.minute_count += 1

    def save_current_minute_average(self, minute_bucket):
        if self.minute_count <= 0:
            return True

        avg_pm10 = int(round(self.minute_sum["pm10"] / self.minute_count))
        avg_pm25 = int(round(self.minute_sum["pm25"] / self.minute_count))
        avg_pm1 = int(round(self.minute_sum["pm1"] / self.minute_count))

        measured_at = minute_bucket.strftime("%Y-%m-%d %H:%M:00")
        batch_data = [(measured_at, self.sensor_index, self.port_name, avg_pm10, avg_pm25, avg_pm1, "NORMAL")]

        success = self.logger.save_measurements_batch(batch_data)
        if success:
            self.clear_minute_buffers()
            return True
        return False

    def check_minute_boundary(self):
        new_bucket = self._get_current_minute()
        if new_bucket == self.current_minute_bucket:
            return

        # 이전 분에 쌓인 데이터가 있다면 저장 시도
        if self.minute_count > 0:
            save_success = self.save_current_minute_average(self.current_minute_bucket)
            if not save_success:
                self.logger.write_error("이전 분 데이터 저장 실패 - 재시도 필요")
        
        # 분이 바뀔 때마다 버킷 갱신 및 카운터 초기화
        self.current_minute_bucket = new_bucket
        self.clear_minute_buffers()

    def get_ui_averages(self):
        pm10_buffer = self.data_buffer["pm10"]
        pm25_buffer = self.data_buffer["pm25"]
        pm1_buffer = self.data_buffer["pm1"]

        if not pm10_buffer or not pm25_buffer or not pm1_buffer:
            return None

        return (
            int(sum(pm10_buffer) / len(pm10_buffer)),
            int(sum(pm25_buffer) / len(pm25_buffer)),
            int(sum(pm1_buffer) / len(pm1_buffer)),
        )

    def run(self):
        is_connected = False
        no_data_error_sent = False
        out_of_range_error_sent = False
        last_ui_update_time = 0.0
        last_data_time = time.time()

        while not self.isInterruptionRequested():
            ser = None
            try:
                ser = serial.Serial(self.port_name, Config.BAUD_RATE, timeout=Config.SERIAL_TIMEOUT)
                ser.reset_input_buffer()

                if not is_connected:
                    self.logger.write_error(f"통신 연결 성공 ({self.port_name})")

                is_connected = True
                no_data_error_sent = False
                out_of_range_error_sent = False
                last_data_time = time.time()
                self.current_minute_bucket = self._get_current_minute()
                self.clear_ui_buffers()

                while ser.is_open and not self.isInterruptionRequested():
                    parsed_values = []
                    while ser.in_waiting > 0:
                        raw_data = ser.readline().decode("utf-8", errors="ignore").strip()
                        if not raw_data:
                            continue

                        parsed = DustParser.parse(raw_data, self.calib_params)
                        if parsed == "OUT_OF_RANGE":
                            if not out_of_range_error_sent:
                                self.error_signal.emit(self.sensor_index, "측정값 범위 초과")
                                self.logger.write_error(f"범위 초과 Raw: {raw_data}")
                                out_of_range_error_sent = True
                            continue

                        if parsed is not None:
                            parsed_values.append(parsed)
                            last_data_time = time.time()
                            out_of_range_error_sent = False

                    self.check_minute_boundary()

                    if time.time() - last_data_time >= Config.NO_DATA_TIMEOUT:
                        if not no_data_error_sent:
                            self.error_signal.emit(self.sensor_index, "데이터 없음")
                            self.logger.write_error(f"{Config.NO_DATA_TIMEOUT:.0f}초 이상 데이터 없음")
                            no_data_error_sent = True
                            self.clear_ui_buffers()
                    elif parsed_values:
                        no_data_error_sent = False
                        for parsed in parsed_values:
                            self.add_measurement(parsed["pm10"], parsed["pm25"], parsed["pm1"])

                        current_time = time.time()
                        if current_time - last_ui_update_time >= 1.0:
                            last_ui_update_time = current_time
                            averages = self.get_ui_averages()
                            if averages is not None:
                                self.data_signal.emit(self.sensor_index, averages[0], averages[1], averages[2])

                    self.safe_sleep(100)

            except serial.SerialException as e:
                if is_connected:
                    self.logger.write_error(f"시리얼 통신 오류: {e}")
                is_connected = False
                self.clear_all_buffers()
                self.error_signal.emit(self.sensor_index, "연결 실패/끊김")
            except Exception:
                if is_connected:
                    self.logger.write_error("시스템 오류:\n" + traceback.format_exc())
                is_connected = False
                self.clear_all_buffers()
                self.error_signal.emit(self.sensor_index, "시스템 오류")
            finally:
                if ser is not None:
                    try:
                        if ser.is_open:
                            ser.close()
                    except Exception:
                        pass
                if not self.isInterruptionRequested():
                    self.safe_sleep(Config.RECONNECT_DELAY_MS)

        try:
            self.save_current_minute_average(self.current_minute_bucket)
        except Exception as e:
            self.logger.write_error(f"Thread 종료 시 마지막 데이터 저장 오류: {e}")


class DustLevelWidget(QWidget):
    def __init__(self, title="미세먼지", levels=None, parent=None):
        super().__init__(parent)
        self.title = title
        self.levels = levels if levels else Config.PM10_LEVELS
        self.current_value = 0
        self.initUI()

    def initUI(self):
        self.setStyleSheet("background-color: transparent; border: none;")
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(3)

        self.title_label = QLabel(self.title)
        self.title_label.setFont(QFont("Malgun Gothic", 11, QFont.Bold))
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("color: black; border: none;")
        main_layout.addWidget(self.title_label)

        value_layout = QHBoxLayout()
        value_layout.setAlignment(Qt.AlignCenter)

        self.val_label = QLabel("----")
        self.val_label.setFont(QFont("Arial", 28, QFont.Bold))
        self.val_label.setStyleSheet("color: black; border: none;")
        value_layout.addWidget(self.val_label)

        unit_label = QLabel("μg/m³")
        unit_label.setFont(QFont("Arial", 10, QFont.Bold))
        unit_label.setStyleSheet("color: #666666; margin-bottom: 4px; border: none;")
        unit_label.setAlignment(Qt.AlignBottom)
        value_layout.addWidget(unit_label)

        main_layout.addLayout(value_layout)

        self.arrow_container = QWidget()
        self.arrow_container.setStyleSheet("border: none;")
        self.arrow_container.setFixedHeight(14)

        self.arrow_label = QLabel("▼")
        self.arrow_label.setFont(QFont("Arial", 9, QFont.Bold))
        self.arrow_label.setStyleSheet("color: black; border: none;")
        self.arrow_label.setAlignment(Qt.AlignCenter)
        self.arrow_label.setFixedWidth(14)
        self.arrow_label.setParent(self.arrow_container)

        main_layout.addWidget(self.arrow_container)

        level_bar_frame = QFrame()
        level_bar_frame.setStyleSheet("border: none;")
        self.level_bar_layout = QHBoxLayout(level_bar_frame)
        self.level_bar_layout.setContentsMargins(0, 0, 0, 0)
        self.level_bar_layout.setSpacing(2)

        self.level_bars = []
        for i, level in enumerate(self.levels):
            bar = QFrame()
            bar.setFixedHeight(10)
            border_radius = ""
            if i == 0:
                border_radius = "border-top-left-radius: 5px; border-bottom-left-radius: 5px;"
            elif i == len(self.levels) - 1:
                border_radius = "border-top-right-radius: 5px; border-bottom-right-radius: 5px;"

            bar.setStyleSheet(f"background-color: {level['color']}; {border_radius}")
            self.level_bars.append(bar)
            self.level_bar_layout.addWidget(bar)

        main_layout.addWidget(level_bar_frame)

        self.status_text_label = QLabel("대기 중...")
        self.status_text_label.setFont(QFont("Malgun Gothic", 11, QFont.Bold))
        self.status_text_label.setAlignment(Qt.AlignCenter)
        self.status_text_label.setFixedHeight(30)
        self.status_text_label.setStyleSheet("background-color: #E0E0E0; border-radius: 6px; color: gray;")
        main_layout.addWidget(self.status_text_label)

    def update_val(self, value):
        self.arrow_label.show()
        self.current_value = value
        self.val_label.setText(str(value))

        current_level = self.levels[self._find_level_index(value)]
        self.status_text_label.setText(current_level["name"])

        text_color = "black" if current_level["name"] == "보통" else "white"
        self.status_text_label.setStyleSheet(
            f"background-color: {current_level['color']}; border-radius: 6px; color: {text_color}; border: none;"
        )
        self.update_arrow_position()

    def _find_level_index(self, value):
        for i, level in enumerate(self.levels):
            if value <= level["max"]:
                return i
        return len(self.levels) - 1

    def set_error_state(self, msg):
        self.val_label.setText("----")
        self.status_text_label.setText(msg)
        self.status_text_label.setStyleSheet("background-color: #FFCDD2; border-radius: 6px; color: #B71C1C; border: none;")
        self.current_value = self.levels[0]["min"]
        self.update_arrow_position()
        self.arrow_label.hide()

    def update_arrow_position(self):
        if not self.level_bars or self.level_bars[0].geometry().width() == 0:
            return

        value = self.current_value
        level_index = self._find_level_index(value)
        current_level = self.levels[level_index]
        level_min, level_max = current_level["min"], current_level["max"]

        if value <= level_min:
            ratio = 0.0
        elif value >= level_max:
            ratio = 1.0
        else:
            denominator = level_max - level_min
            ratio = 0.0 if denominator <= 0 else (value - level_min) / denominator

        target_bar = self.level_bars[level_index]
        target_x = target_bar.geometry().x() + (target_bar.geometry().width() * ratio)
        arrow_x = int(target_x - (self.arrow_label.width() / 2))
        max_x = self.arrow_container.width() - self.arrow_label.width()
        arrow_x = max(0, min(arrow_x, max_x))

        self.arrow_label.move(QPoint(arrow_x, int(self.arrow_container.height() - self.arrow_label.height() + 2)))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_arrow_position()


class PortSelectionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected_ports = {}
        self.available_ports = sorted([port.device for port in serial.tools.list_ports.comports()])
        self.initUI()

    def initUI(self):
        self.setWindowTitle("센서 포트 설정")
        self.resize(350, 250)
        self.setStyleSheet("background-color: white;")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        title_label = QLabel("모니터링할 센서 포트를 선택하세요")
        title_label.setFont(QFont("Malgun Gothic", 10, QFont.Bold))
        layout.addWidget(title_label)

        self.combos = []
        for i in range(Config.MAX_SENSORS):
            h_layout = QHBoxLayout()
            label = QLabel(f"센서 {i + 1} 포트:")
            label.setFont(QFont("Malgun Gothic", 9))

            combo = QComboBox()
            combo.addItem("선택 안 함")
            combo.addItems(self.available_ports)
            combo.setFont(QFont("Malgun Gothic", 9))
            combo.currentIndexChanged.connect(self.update_combo_items)

            h_layout.addWidget(label)
            h_layout.addWidget(combo)
            layout.addLayout(h_layout)
            self.combos.append(combo)

        self.start_btn = QPushButton("모니터링 시작")
        self.start_btn.setFixedHeight(35)
        self.start_btn.setFont(QFont("Malgun Gothic", 10, QFont.Bold))
        self.start_btn.setStyleSheet("background-color: #007BFF; color: white; border-radius: 4px;")
        self.start_btn.clicked.connect(self.accept_selection)
        layout.addWidget(self.start_btn)

    def update_combo_items(self):
        selected_values = [
            combo.currentText() for combo in self.combos
            if combo.currentText() and combo.currentText() != "선택 안 함"
        ]

        for combo in self.combos:
            current_selection = combo.currentText()
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("선택 안 함")

            for port in self.available_ports:
                if port in selected_values and port != current_selection:
                    continue
                combo.addItem(port)

            index = combo.findText(current_selection)
            combo.setCurrentIndex(index if index != -1 else 0)
            combo.blockSignals(False)

    def accept_selection(self):
        self.selected_ports = {
            i: combo.currentText() for i, combo in enumerate(self.combos)
            if combo.currentText() and combo.currentText() != "선택 안 함"
        }
        if not self.selected_ports:
            QMessageBox.warning(self, "경고", "최소 하나 이상의 센서 포트를 선택해 주세요.")
            return
        self.accept()


class DustMonitorApp(QMainWindow):
    def __init__(self, slot_mapping):
        super().__init__()
        self.threads = []
        self.sensor_widgets = {}
        self.slot_mapping = slot_mapping
        self.db_path = os.path.join(BASE_DIR, "Data", "dust_measurement.db")
        self.initUI()
        self.start_monitoring()

    def initUI(self):
        self.setWindowTitle("다중 미세먼지 모니터링 시스템")
        self.setStyleSheet("QMainWindow { background-color: white; }")
        self.setMinimumSize(400, 150)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        self.setCentralWidget(scroll)

        scroll_content = QWidget()
        scroll.setWidget(scroll_content)

        self.scroll_layout = QVBoxLayout(scroll_content)
        self.scroll_layout.setContentsMargins(15, 15, 15, 15)
        self.scroll_layout.setSpacing(15)
        self.scroll_layout.addStretch(1)

    def start_monitoring(self):
        for sensor_index, port_name in self.slot_mapping.items():
            group_box = QGroupBox(f"센서 {sensor_index + 1} 포트: {port_name}")
            group_box.setFont(QFont("Malgun Gothic", 11, QFont.Bold))

            port_layout = QHBoxLayout(group_box)
            port_layout.setSpacing(10)

            w_pm10 = DustLevelWidget("PM10 (미세먼지)", Config.PM10_LEVELS)
            w_pm25 = DustLevelWidget("PM2.5 (초미세먼지)", Config.PM25_LEVELS)
            w_pm1 = DustLevelWidget("PM1.0 (극미세먼지)", Config.PM1_LEVELS)

            port_layout.addWidget(w_pm10)
            port_layout.addWidget(w_pm25)
            port_layout.addWidget(w_pm1)

            self.scroll_layout.insertWidget(self.scroll_layout.count() - 1, group_box)
            self.sensor_widgets[sensor_index] = {"pm10": w_pm10, "pm25": w_pm25, "pm1": w_pm1}

            thread = SerialThread(port_name, sensor_index)
            thread.data_signal.connect(self.update_data)
            thread.error_signal.connect(self.handle_error)
            thread.start()
            self.threads.append(thread)

        sensor_count = len(self.slot_mapping)
        self.resize(980, max(220, sensor_count * 245 + 60))

    def update_data(self, sensor_index, pm10, pm25, pm1):
        widgets = self.sensor_widgets.get(sensor_index)
        if not widgets:
            return
        widgets["pm10"].update_val(pm10)
        widgets["pm25"].update_val(pm25)
        widgets["pm1"].update_val(pm1)

    def handle_error(self, sensor_index, error_msg):
        widgets = self.sensor_widgets.get(sensor_index)
        if not widgets:
            return
        widgets["pm10"].set_error_state(error_msg)
        widgets["pm25"].set_error_state(error_msg)
        widgets["pm1"].set_error_state(error_msg)

    def closeEvent(self, event):
        for thread in self.threads:
            if thread.isRunning():
                thread.requestInterruption()

        for thread in self.threads:
            if thread.isRunning():
                if not thread.wait(Config.THREAD_WAIT_MS):
                    thread.logger.write_error("프로그램 종료 시 Thread 종료 시간 초과")

        for thread in self.threads:
            try:
                if thread.isRunning():
                    continue
                thread.save_current_minute_average(thread.current_minute_bucket)
            except Exception as e:
                thread.logger.write_error(f"종료 데이터 저장 오류: {e}")

        for thread in self.threads:
            try:
                thread.logger.export_excel()
            except Exception as e:
                thread.logger.write_error(f"종료 시 Excel Export 오류: {e}")

        DatabaseManager(self.db_path).stop()
        event.accept()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.dragPos = event.globalPos()

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton and hasattr(self, "dragPos"):
            self.move(self.pos() + event.globalPos() - self.dragPos)
            self.dragPos = event.globalPos()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.close()
        else:
            super().keyPressEvent(event)


def main():
    app = QApplication(sys.argv)
    db_path = os.path.join(BASE_DIR, "Data", "dust_measurement.db")
    SensorLogger.init_global_database(db_path)
    SensorLogger.cleanup_old_logs_global()

    dialog = PortSelectionDialog()
    if dialog.exec_() != QDialog.Accepted:
        return 0

    main_window = DustMonitorApp(dialog.selected_ports)
    main_window.show()
    return app.exec_()


if __name__ == "__main__":
    sys.exit(main())