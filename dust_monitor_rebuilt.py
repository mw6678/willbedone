import csv
import sqlite3
from datetime import datetime
from datetime import datetime
import sys
import os
import serial
import time
import traceback
import serial.tools.list_ports

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QFrame, QPushButton, QDialog, QMessageBox,
    QListWidget, QListWidgetItem, QScrollArea, QGroupBox
)
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QPoint
from PyQt5.QtGui import QFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ==========================================
# 1. 설정 관리 (Config)
# ==========================================
class Config:
    BAUD_RATE = 9600
    # 센서 포트별 [PM10 보정값, PM2.5 보정값, PM1.0 보정값]
    # 센서 포트별 [(PM10 Scale, Offset), (PM2.5 Scale, Offset), (PM1.0 Scale, Offset)]
    # Scale 기본값 = 1.0 (변화 없음), Offset 기본값 = 0.0 즉 최종값 = max(0,(Raw 값 + Offset)*Scale)
    SENSOR_CALIBRATION = {
        0: {"scale": (1.0, 1.0, 1.0), "offset": (0.0, 0.0, 0.0)},
        1: {"scale": (1.0, 1.0, 1.0), "offset": (0.0, 0.0, 0.0)},
        2: {"scale": (1.0, 1.0, 1.0), "offset": (0.0, 0.0, 0.0)},
        3: {"scale": (1.0, 1.0, 1.0), "offset": (0.0, 0.0, 0.0)},
    }
    
    SMOOTHING_WINDOW = 1
    NO_DATA_TIMEOUT = 30.0
    RECONNECT_DELAY_MS = 3000
    LOG_RETENTION_DAYS = 30

    PM10_LEVELS = [
        {"name": "좋음", "min": 0, "max": 30, "color": "#28A745"},
        {"name": "보통", "min": 31, "max": 80, "color": "#FFD700"},
        {"name": "민감군", "min": 81, "max": 120, "color": "#FD7E14"},
        {"name": "나쁨", "min": 121, "max": 150, "color": "#DC3545"},
        {"name": "매우 나쁨", "min": 151, "max": 300, "color": "#800080"},
        {"name": "위험", "min": 301, "max": 600, "color": "#795548"},
    ]
    PM25_LEVELS = [
        {"name": "좋음", "min": 0, "max": 15, "color": "#28A745"},
        {"name": "보통", "min": 16, "max": 35, "color": "#FFD700"},
        {"name": "민감군", "min": 36, "max": 50, "color": "#FD7E14"},
        {"name": "나쁨", "min": 51, "max": 75, "color": "#DC3545"},
        {"name": "매우 나쁨", "min": 76, "max": 100, "color": "#800080"},
        {"name": "위험", "min": 101, "max": 500, "color": "#795548"},
    ]
    PM1_LEVELS = [
        {"name": "좋음", "min": 0, "max": 10, "color": "#28A745"},
        {"name": "보통", "min": 11, "max": 25, "color": "#FFD700"},
        {"name": "민감군", "min": 26, "max": 35, "color": "#FD7E14"},
        {"name": "나쁨", "min": 36, "max": 50, "color": "#DC3545"},
        {"name": "매우 나쁨", "min": 51, "max": 75, "color": "#800080"},
        {"name": "위험", "min": 76, "max": 300, "color": "#795548"},
    ]

# ==========================================
# 2. 데이터 파싱 로직 (Data Parser)
# ==========================================
# ==========================================
# 2. 데이터 파싱 로직 (Data Parser)
# ==========================================
class DustParser:

    @staticmethod
    def parse(raw_data, calib_params):
        try:
            raw_data = raw_data.strip()
            if not raw_data:
                return None

            parts = [
                item.strip() for item in raw_data.replace(" ", "").split(",")
            ]
            if len(parts) < 3:
                return None

            # calib_params = {"scale": (s10, s25, s1), "offset": (o10, o25, o1)}
            scales = calib_params.get("scale", (1.0, 1.0, 1.0))
            offsets = calib_params.get("offset", (0.0, 0.0, 0.0))

            scale_pm10, scale_pm25, scale_pm1 = scales
            offset_pm10, offset_pm25, offset_pm1 = offsets

            # Raw 데이터 읽기
            raw_pm10 = float(parts[2])
            raw_pm25 = float(parts[1])
            raw_pm1 = float(parts[0])

            # 비율 보정 및 오프셋 적용 공식: max(0, int(round((raw + offset) * scale)))
            pm10 = max(0, int(round((raw_pm10 + offset_pm10) * scale_pm10)))
            pm25 = max(0, int(round((raw_pm25 + offset_pm25) * scale_pm25)))
            pm1 = max(0, int(round((raw_pm1 + offset_pm1) * scale_pm1)))

            if pm1 > 1000 or pm25 > 1000 or pm10 > 2000:
                return "OUT_OF_RANGE"

            return {
                "raw": {"pm10": raw_pm10, "pm25": raw_pm25, "pm1": raw_pm1,
                        },
                "value": {"pm10": pm10, "pm25": pm25, "pm1": pm1,
                          }
                }
        except Exception:
            return None

# ==========================================
# 3. 로깅 및 파일 처리 (Sensor Logger)
# ==========================================
class SensorLogger:
    def __init__(self, port_name, sensor_index):
        self.port_name = port_name
        self.sensor_index = sensor_index
        
        self.log_dir = os.path.join(BASE_DIR, "Logs")
        self.excel_dir = os.path.join(BASE_DIR, "Excel_Logs") # CSV 대신 엑셀 폴더
        self.db_dir = os.path.join(BASE_DIR, "Data")
        
        os.makedirs(self.log_dir, exist_ok=True)
        os.makedirs(self.excel_dir, exist_ok=True)
        os.makedirs(self.db_dir, exist_ok=True)

        self.db_path = os.path.join(self.db_dir, "dust_measurement.db")
        # ⚠️ init_database()와 cleanup_old_logs()는 여기서 호출하지 않고 외부(메인)로 뺍니다!

    # [클래스 외부 혹은 전역 함수로 분리할 수도 있는 초기화/정리 함수들]
    @staticmethod
    def init_global_database(db_path):
        """앱 시작 시 전체 단 한 번만 호출하여 DB 초기화 및 최적화"""
        try:
            with sqlite3.connect(db_path) as conn:
                conn.execute("PRAGMA journal_mode=WAL;")
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS measurements (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        measured_at TEXT NOT NULL,
                        sensor_index INTEGER NOT NULL,
                        port TEXT NOT NULL,
                        raw_pm10 REAL,
                        raw_pm25 REAL,
                        raw_pm1 REAL,
                        pm10 REAL,
                        pm25 REAL,
                        pm1 REAL,
                        status TEXT DEFAULT 'NORMAL'
                    )
                """)
                conn.execute("CREATE INDEX IF NOT EXISTS idx_measurements_time ON measurements(measured_at)")
                conn.execute("CREATE INDEX IF NOT EXISTS idx_measurements_sensor ON measurements(sensor_index)")
                conn.commit()
        except Exception as e:
            print(f"Global SQLite 초기화 오류: {e}")

    @staticmethod
    def cleanup_old_logs_global():
        """앱 시작 시 오래된 로그 파일 정리"""
        threshold_time = time.time() - (Config.LOG_RETENTION_DAYS * 24 * 60 * 60)
        for target_dir in [os.path.join(BASE_DIR, "Logs"), os.path.join(BASE_DIR, "Excel_Logs")]:
            if not os.path.exists(target_dir):
                continue
            for filename in os.listdir(target_dir):
                file_path = os.path.join(target_dir, filename)
                if os.path.isfile(file_path):
                    try:
                        if os.path.getmtime(file_path) < threshold_time:
                            os.remove(file_path)
                    except Exception:
                        pass

    def write_error(self, message):
        log_file = os.path.join(self.log_dir, f"error_log_Sensor{self.sensor_index + 1}.txt")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with open(log_file, "a", encoding="utf-8-sig") as f:
                f.write(f"{timestamp} | 포트: {self.port_name} | {message}\n")
        except Exception:
            pass

    def save_measurement(self, measured_at, raw_pm10, raw_pm25, raw_pm1, pm10, pm25, pm1, status="NORMAL"):
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.execute("""
                    INSERT INTO measurements (
                        measured_at, sensor_index, port,
                        raw_pm10, raw_pm25, raw_pm1,
                        pm10, pm25, pm1, status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (measured_at, self.sensor_index, self.port_name, raw_pm10, raw_pm25, raw_pm1, pm10, pm25, pm1, status))
                conn.commit()
            return True
        except Exception as e:
            self.write_error(f"SQLite 측정 데이터 저장 오류: {e}")
            return False

    def export_excel(self, date_str=None):
       
        try:
            with sqlite3.connect(self.db_path) as conn:
                if date_str:
                    rows = conn.execute("""
                        SELECT measured_at, port, raw_pm10, raw_pm25, raw_pm1, pm10, pm25, pm1, status
                        FROM measurements WHERE sensor_index = ? AND measured_at LIKE ? ORDER BY measured_at
                    """, (self.sensor_index, f"{date_str}%")).fetchall()
                else:
                    rows = conn.execute("""
                        SELECT measured_at, port, raw_pm10, raw_pm25, raw_pm1, pm10, pm25, pm1, status
                        FROM measurements WHERE sensor_index = ? ORDER BY measured_at
                    """, (self.sensor_index,)).fetchall()

            if not rows:
                return False

            filename = f"Dust_log_Sensor{self.sensor_index + 1}.xlsx"
            file_path = os.path.join(self.excel_dir, filename)

            wb = Workbook()
            ws = wb.active
            ws.title = "측정 데이터"

            headers = ["측정일시", "포트", "Raw_PM10", "Raw_PM2.5", "Raw_PM1.0", "PM10", "PM2.5", "PM1.0", "상태"]
            ws.append(headers)
            ws.row_dimensions[1].height = 25

            # 스타일 정의
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
            )
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
            data_font = Font(name="Malgun Gothic", size=9)
            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            for row_data in rows:
                ws.append(list(row_data))
                current_row = ws.max_row
                ws.row_dimensions[current_row].height = 18

                for col_num, val in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.font = data_font
                    cell.border = thin_border

                    # 정렬 및 포맷팅
                    if col_num in [1, 2, 9]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = right_align
                        if isinstance(val, (int, float)):
                            cell.number_format = '#,##0'

            # 열 너비 자동 조절
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    max_len = max(max_len, len(val_str.encode('utf-8')) if cell.row == 1 else len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            wb.save(file_path)
            return True

        except Exception as e:
            self.write_error(f"Excel Export 오류: {e}")
            return False
            
# ==========================================
# SQLite 초기화
# ==========================================
    def init_database(self):
        try:
            with sqlite3.connect(self.db_path) as conn:
                # 동시성 성능을 크게 높여주는 WAL 모드 활성화
                conn.execute("PRAGMA journal_mode=WAL;")
                
                conn.execute("""
                    CREATE TABLE IF NOT EXISTS measurements (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        measured_at TEXT NOT NULL,
                        sensor_index INTEGER NOT NULL,
                        port TEXT NOT NULL,
                        raw_pm10 REAL,
                        raw_pm25 REAL,
                        raw_pm1 REAL,
                        pm10 REAL,
                        pm25 REAL,
                        pm1 REAL,
                        status TEXT DEFAULT 'NORMAL'
                    )
                """)

                conn.execute("""
                    CREATE INDEX IF NOT EXISTS idx_measurements_time
                    ON measurements(measured_at)
                """)

                conn.execute("""
                    CREATE INDEX IF NOT EXISTS idx_measurements_sensor
                    ON measurements(sensor_index)
                """)

                conn.commit()

        except Exception as e:
            self.write_error(f"SQLite 초기화 오류: {e}")
    # ==========================================
    # 오류 로그
    # ==========================================

    def write_error(self, message):
        log_file = os.path.join(self.log_dir, f"error_log_Sensor{self.sensor_index + 1}.txt")
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with open(log_file, "a", encoding="utf-8-sig") as f:
                f.write(f"{timestamp} | 포트: {self.port_name} | {message}\n")
        except Exception:
            pass

    # ==========================================
    # 측정 데이터 SQLite 저장
    # ==========================================
    def save_measurement(
        self,
        measured_at,
        raw_pm10,
        raw_pm25,
        raw_pm1,
        pm10,
        pm25,
        pm1,
        status="NORMAL"
    ):
        try:
            with sqlite3.connect(self.db_path) as conn:

                conn.execute("""
                    INSERT INTO measurements (
                        measured_at,
                        sensor_index,
                        port,

                        raw_pm10,
                        raw_pm25,
                        raw_pm1,

                        pm10,
                        pm25,
                        pm1,

                        status
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    measured_at,
                    self.sensor_index,
                    self.port_name,

                    raw_pm10,
                    raw_pm25,
                    raw_pm1,

                    pm10,
                    pm25,
                    pm1,

                    status
                ))

                conn.commit()

            return True

        except Exception as e:
            self.write_error(
                f"SQLite 측정 데이터 저장 오류: {e}"
            )
            return False

    # ==========================================
    # SQLite → CSV Export
    # ==========================================
    def export_csv(self, date_str=None):

        try:
            with sqlite3.connect(self.db_path) as conn:

                if date_str:
                    rows = conn.execute("""
                        SELECT
                            measured_at,
                            port,
                            raw_pm10,
                            raw_pm25,
                            raw_pm1,
                            pm10,
                            pm25,
                            pm1,
                            status
                        FROM measurements
                        WHERE measured_at LIKE ?
                        ORDER BY measured_at
                    """, (
                        f"{date_str}%",
                    )).fetchall()

                else:
                    rows = conn.execute("""
                        SELECT
                            measured_at,
                            port,
                            raw_pm10,
                            raw_pm25,
                            raw_pm1,
                            pm10,
                            pm25,
                            pm1,
                            status
                        FROM measurements
                        ORDER BY measured_at
                    """).fetchall()

            if date_str:
                filename = (
                    f"Dust_log_{date_str}"
                    f"_Sensor{self.sensor_index + 1}.csv"
                )
            else:
                filename = (
                    f"Dust_log_Sensor"
                    f"{self.sensor_index + 1}.csv"
                )

            file_path = os.path.join(
                self.csv_dir,
                filename
            )

            with open(
                file_path,
                "w",
                newline="",
                encoding="utf-8-sig"
            ) as f:

                writer = csv.writer(f)

                writer.writerow([
                    "측정일시",
                    "포트",

                    "Raw_PM10",
                    "Raw_PM2.5",
                    "Raw_PM1.0",

                    "PM10",
                    "PM2.5",
                    "PM1.0",

                    "상태"
                ])

                writer.writerows(rows)

            return True

        except Exception as e:
            self.write_error(
                f"CSV Export 오류: {e}"
            )
            return False

    # ==========================================
    # 오래된 로그 정리
    # ==========================================
    def cleanup_old_logs(self):

        threshold_time = (
            time.time()
            - (
                Config.LOG_RETENTION_DAYS
                * 24
                * 60
                * 60
            )
        )

        for target_dir in [
            self.log_dir,
            self.csv_dir
        ]:

            if not os.path.exists(target_dir):
                continue

            for filename in os.listdir(target_dir):

                file_path = os.path.join(
                    target_dir,
                    filename
                )

                if os.path.isfile(file_path):

                    try:

                        if (
                            os.path.getmtime(file_path)
                            < threshold_time
                        ):
                            os.remove(file_path)

                    except Exception:
                        pass

# ==========================================
# 4. 시리얼 통신 스레드 (Serial Thread)
# ==========================================
class SerialThread(QThread):
    data_signal = pyqtSignal(str, int, int, int)
    error_signal = pyqtSignal(str, str)

    def __init__(self, port_name, sensor_index):
        super().__init__()
        self.port_name = port_name

        # scale과 offset을 담은 딕셔너리를 통째로 전달
        default_calib = {"scale": (1.0, 1.0, 1.0), "offset": (0.0, 0.0, 0.0)}
        self.calib_params = Config.SENSOR_CALIBRATION.get(
            sensor_index, default_calib
        )

        self.logger = SensorLogger(port_name, sensor_index)
        self.data_buffer = {"pm10": [], "pm25": [], "pm1": []}
        self.minute_data_buffer = {"pm10": [], "pm25": [], "pm1": []}

    def safe_sleep(self, ms):
        elapsed = 0
        while elapsed < ms:
            if self.isInterruptionRequested():
                return
            sleep_time = min(100, ms - elapsed)
            QThread.msleep(sleep_time)
            elapsed += sleep_time

    def clear_buffers(self):
        for key in self.data_buffer:
            self.data_buffer[key].clear()
            self.minute_data_buffer[key].clear()

    def run(self):
        is_connected = False
        no_data_error_sent = False
        last_saved_minute_key = datetime.now().strftime("%Y-%m-%d %H:%M")
        last_ui_update_time = 0.0
        last_data_time = time.time()

        while not self.isInterruptionRequested():
            ser = None
            try:
                ser = serial.Serial(self.port_name, Config.BAUD_RATE, timeout=0.1)
                ser.reset_input_buffer()

                if not is_connected:
                    self.logger.write_error(f"통신 연결 성공 ({self.port_name})")

                is_connected = True
                no_data_error_sent = False
                last_data_time = time.time()

                while ser.is_open and not self.isInterruptionRequested():
                    # 1. 데이터 수신 및 파싱
                    parsed_values = []
                    while ser.in_waiting > 0:
                        raw_data = ser.readline().decode("utf-8", errors="ignore").strip()
                        if not raw_data:
                            continue

                        # self.offsets -> self.calib_params 로 수정
                        parsed = DustParser.parse(raw_data, self.calib_params)
                        if parsed == "OUT_OF_RANGE":
                            self.error_signal.emit(self.port_name, "측정값 범위 초과")
                            self.logger.write_error(f"범위 초과 Raw: {raw_data}")
                            continue
                        elif parsed is not None:
                            parsed_values.append(parsed)
                            last_data_time = time.time()

                    # 2. 타임아웃 체크
                    if time.time() - last_data_time >= Config.NO_DATA_TIMEOUT:
                        if not no_data_error_sent:
                            self.error_signal.emit(self.port_name, "데이터 없음")
                            self.logger.write_error("30초 이상 데이터 없음")
                            no_data_error_sent = True
                            self.clear_buffers()

                    # 3. 정상 데이터 버퍼링 및 UI 업데이트
                    elif parsed_values:
                        no_data_error_sent = False

                        for parsed in parsed_values:
                            raw = parsed["raw"]
                            value = parsed["value"]
                            pm10 = value["pm10"]
                            pm25 = value["pm25"]
                            pm1 = value["pm1"]

    # --------------------------------
    # SQLite 원본 측정 데이터 저장
    # --------------------------------

                            self.logger.save_measurement(
                                measured_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"),
                                raw_pm10=raw["pm10"],
                                raw_pm25=raw["pm25"],
                                raw_pm1=raw["pm1"],
                                pm10=pm10,
                                pm25=pm25,
                                pm1=pm1,
                                status="NORMAL")
                            
                            for key, val in zip(["pm10", "pm25", "pm1"], (pm10, pm25, pm1)):
                                self.minute_data_buffer[key].append(val)
                                self.data_buffer[key].append(val)
                                if len(self.data_buffer[key]) > Config.SMOOTHING_WINDOW:
                                    self.data_buffer[key].pop(0)

                        if time.time() - last_ui_update_time >= 1.0:
                            last_ui_update_time = time.time()
                            avgs = [
                                int(sum(self.data_buffer[k]) / len(self.data_buffer[k]))
                                for k in ["pm10", "pm25", "pm1"]
                            ]
                            self.data_signal.emit(self.port_name, *avgs)

                    self.safe_sleep(100)

            except serial.SerialException as e:
                if is_connected:
                    self.logger.write_error(f"시리얼 통신 오류: {e}")
                is_connected = False
                self.clear_buffers()
                self.error_signal.emit(self.port_name, "연결 실패/끊김")
            except Exception:
                if is_connected:
                    self.logger.write_error("시스템 오류:\n" + traceback.format_exc())
                is_connected = False
                self.clear_buffers()
                self.error_signal.emit(self.port_name, "시스템 오류")
            finally:
                if ser is not None:
                    try:
                        if ser.is_open:
                            ser.close()
                    except Exception:
                        pass
                if not self.isInterruptionRequested():
                    self.safe_sleep(Config.RECONNECT_DELAY_MS)
                    


# ==========================================
# 5. UI 위젯 및 앱 메인 로직 (기존과 동일)
# ==========================================
class PortSelectionDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("미세먼지 센서 포트 선택")
        self.resize(350, 400)
        self.setStyleSheet("background-color: white;")
        self.selected_ports = []
        self.initUI()
        self.refresh_ports()

    def initUI(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        title = QLabel("연결할 COM 포트를 모두 선택하세요")
        title.setFont(QFont("Malgun Gothic", 11, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        self.list_widget = QListWidget()
        layout.addWidget(self.list_widget)

        refresh_btn = QPushButton("포트 새로고침")
        refresh_btn.clicked.connect(self.refresh_ports)
        layout.addWidget(refresh_btn)

        self.start_btn = QPushButton("모니터링 시작")
        self.start_btn.setFixedHeight(40)
        self.start_btn.setFont(QFont("Malgun Gothic", 10, QFont.Bold))
        self.start_btn.setStyleSheet("background-color: #007BFF; color: white; border-radius: 5px;")
        self.start_btn.clicked.connect(self.on_start_clicked)
        layout.addWidget(self.start_btn)

    def refresh_ports(self):
        self.list_widget.clear()
        ports = [p.device for p in serial.tools.list_ports.comports()]
        if not ports:
            item = QListWidgetItem("연결된 포트 없음")
            item.setFlags(Qt.NoItemFlags)
            self.list_widget.addItem(item)
            return

        for port in ports:
            item = QListWidgetItem(port)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            self.list_widget.addItem(item)

    def on_start_clicked(self):
        self.selected_ports = [self.list_widget.item(i).text() for i in range(self.list_widget.count()) if self.list_widget.item(i).checkState() == Qt.Checked]
        if not self.selected_ports:
            QMessageBox.warning(self, "경고", "최소 하나 이상의 포트를 선택해 주세요.")
            return
        if len(self.selected_ports) != len(set(self.selected_ports)):
            QMessageBox.warning(self, "경고", "같은 COM 포트를 중복 선택할 수 없습니다.")
            return
        self.accept()

class DustLevelWidget(QWidget):
    def __init__(self, title="미세먼지", levels=None, parent=None):
        super().__init__(parent)
        self.title = title
        self.levels = levels if levels else Config.PM10_LEVELS
        self.current_value = 0
        self.initUI()

    def initUI(self):
        self.setStyleSheet("background-color: transparent; border: none;")        
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(3)

        self.title_label = QLabel(self.title)
        self.title_label.setFont(QFont("Malgun Gothic", 11, QFont.Bold))
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_label.setStyleSheet("color: black; border: none;")
        main_layout.addWidget(self.title_label)

        value_layout = QHBoxLayout()
        value_layout.setAlignment(Qt.AlignCenter)

        self.val_label = QLabel("----")
        self.val_label.setFont(QFont("Arial", 28, QFont.Bold))
        self.val_label.setStyleSheet("color: black; border: none;")
        value_layout.addWidget(self.val_label)

        unit_label = QLabel("µg/m³")
        unit_label.setFont(QFont("Arial", 10, QFont.Bold))
        unit_label.setStyleSheet("color: #666666; margin-bottom: 4px; border: none;")
        unit_label.setAlignment(Qt.AlignBottom)
        value_layout.addWidget(unit_label)

        main_layout.addLayout(value_layout)

        self.arrow_container = QWidget()
        self.arrow_container.setStyleSheet("border: none;")
        self.arrow_container.setFixedHeight(14)

        self.arrow_label = QLabel("▼")
        self.arrow_label.setFont(QFont("Arial", 9, QFont.Bold))
        self.arrow_label.setStyleSheet("color: black; border: none;")
        self.arrow_label.setAlignment(Qt.AlignCenter)
        self.arrow_label.setFixedWidth(14)
        self.arrow_label.setParent(self.arrow_container)
        main_layout.addWidget(self.arrow_container)

        level_bar_frame = QFrame()
        level_bar_frame.setStyleSheet("border: none;")
        self.level_bar_layout = QHBoxLayout(level_bar_frame)
        self.level_bar_layout.setContentsMargins(0, 0, 0, 0)
        self.level_bar_layout.setSpacing(2)

        self.level_bars = []
        for i, level in enumerate(self.levels):
            bar = QFrame()
            bar.setFixedHeight(10)
            border_radius = ""
            if i == 0: border_radius = "border-top-left-radius: 5px; border-bottom-left-radius: 5px;"
            elif i == len(self.levels) - 1: border_radius = "border-top-right-radius: 5px; border-bottom-right-radius: 5px;"
            bar.setStyleSheet(f"background-color: {level['color']}; {border_radius}")
            self.level_bars.append(bar)
            self.level_bar_layout.addWidget(bar)
        main_layout.addWidget(level_bar_frame)

        self.status_text_label = QLabel("대기 중...")
        self.status_text_label.setFont(QFont("Malgun Gothic", 11, QFont.Bold))
        self.status_text_label.setAlignment(Qt.AlignCenter)
        self.status_text_label.setFixedHeight(30)
        self.status_text_label.setStyleSheet("background-color: #E0E0E0; border-radius: 6px; color: gray;")
        main_layout.addWidget(self.status_text_label)

    def update_val(self, value):
        self.arrow_label.show()
        self.current_value = value
        self.val_label.setText(str(value))

        current_level = self.levels[self._find_level_index(value)]

        self.status_text_label.setText(current_level["name"])
        text_color = "black" if current_level["name"] == "보통" else "white"
        self.status_text_label.setStyleSheet(f"background-color: {current_level['color']}; border-radius: 6px; color: {text_color}; border: none;")
        self.update_arrow_position()

    def _find_level_index(self, value):
        for i, level in enumerate(self.levels):
            if value <= level["max"]:
                return i
        return len(self.levels) - 1

    def set_error_state(self, msg):
        self.val_label.setText("----")
        self.status_text_label.setText(msg)
        self.status_text_label.setStyleSheet("background-color: #FFCDD2; border-radius: 6px; color: #B71C1C; border: none;")
        self.current_value = self.levels[0]["min"]
        self.update_arrow_position()
        self.arrow_label.hide()

    def update_arrow_position(self):
        if not self.level_bars or self.level_bars[0].geometry().width() == 0:
            return

        value = self.current_value
        level_index = self._find_level_index(value)

        current_level = self.levels[level_index]
        level_min, level_max = current_level["min"], current_level["max"]

        if value <= level_min: ratio = 0.0
        elif value >= level_max: ratio = 1.0
        else: ratio = (value - level_min) / (level_max - level_min)

        target_bar = self.level_bars[level_index]
        target_x = target_bar.geometry().x() + (target_bar.geometry().width() * ratio)
        # 1. 화살표가 위치할 X 좌표 계산
        arrow_x = int(target_x - (self.arrow_label.width() / 2))
        # 2. 화살표가 잘리지 않도록 최소 0, 최대 (컨테이너 너비 - 화살표 너비) 사이로 제한
        max_x = self.arrow_container.width() - self.arrow_label.width()
        arrow_x = max(0, min(arrow_x, max_x))
        # 3. 제한된 좌표를 적용하여 화살표 이동
        self.arrow_label.move(QPoint(arrow_x, int(self.arrow_container.height() - self.arrow_label.height() + 2)))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.update_arrow_position()

class DustMonitorApp(QMainWindow):
    def __init__(self, ports):
        super().__init__()
        self.ports = ports
        self.threads = []
        self.port_widgets = {}
        self.initUI()
        self.start_threads()

    def initUI(self):
        self.setWindowTitle("다중 미세먼지 모니터링 시스템")
        self.resize(1000, 700)
        self.setStyleSheet("QMainWindow { background-color: white; }")

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        self.setCentralWidget(scroll)

        container = QWidget()
        scroll.setWidget(container)
        main_layout = QVBoxLayout(container)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)

        for port in self.ports:
            group_box = QGroupBox(f"센서 포트: {port}")
            group_box.setFont(QFont("Malgun Gothic", 11, QFont.Bold))
            port_layout = QHBoxLayout(group_box)
            port_layout.setSpacing(10)

            w_pm10 = DustLevelWidget("PM10 (미세먼지)", Config.PM10_LEVELS)
            w_pm25 = DustLevelWidget("PM2.5 (초미세먼지)", Config.PM25_LEVELS)
            w_pm1 = DustLevelWidget("PM1.0 (극미세먼지)", Config.PM1_LEVELS)

            port_layout.addWidget(w_pm10)
            port_layout.addWidget(w_pm25)
            port_layout.addWidget(w_pm1)
            main_layout.addWidget(group_box)

            self.port_widgets[port] = {"pm10": w_pm10, "pm25": w_pm25, "pm1": w_pm1}

        main_layout.addStretch(1)

    def start_threads(self):
        for index, port in enumerate(self.ports):
            thread = SerialThread(port, sensor_index=index)
            thread.data_signal.connect(self.update_data)
            thread.error_signal.connect(self.handle_error)
            self.threads.append(thread)
            thread.start()

    def update_data(self, port, pm10, pm25, pm1):
        widgets = self.port_widgets.get(port)
        if widgets:
            widgets["pm10"].update_val(pm10)
            widgets["pm25"].update_val(pm25)
            widgets["pm1"].update_val(pm1)

    def handle_error(self, port, error_msg):
        widgets = self.port_widgets.get(port)
        if widgets:
            widgets["pm10"].set_error_state(error_msg)
            widgets["pm25"].set_error_state(error_msg)
            widgets["pm1"].set_error_state(error_msg)

    def closeEvent(self, event):
        for thread in self.threads:
            thread.requestInterruption()
        
        for thread in self.threads:
            thread.wait(2000)
    
        # [변경] 종료할 때 각 센서별로 엑셀 파일 내보내기 실행
        for thread in self.threads:
            thread.logger.export_excel()
            
        event.accept()
            
        event.accept()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton: self.dragPos = event.globalPos()

    def mouseMoveEvent(self, event):
        if event.buttons() == Qt.LeftButton and hasattr(self, "dragPos"):
            self.move(self.pos() + event.globalPos() - self.dragPos)
            self.dragPos = event.globalPos()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape: self.close()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    # [추가] 앱 시작 시 전역 DB 초기화 및 오래된 로그 정리 (중복 실행 방지)
    db_path = os.path.join(BASE_DIR, "Data", "dust_measurement.db")
    SensorLogger.init_global_database(db_path)
    SensorLogger.cleanup_old_logs_global()

    setup_dialog = PortSelectionDialog()
    if setup_dialog.exec_() == QDialog.Accepted:
        window = DustMonitorApp(setup_dialog.selected_ports)
        window.show()
        sys.exit(app.exec_())
